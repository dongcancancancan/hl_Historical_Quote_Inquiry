import sys
from openpyxl import load_workbook

file_path = "data/original_excels/171a94d36b5c480a8c3b972d6e73cea2.xlsx"

try:
    print(f"Loading workbook: {file_path}")
    wb = load_workbook(file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
            # 截取前10列，将 None 转为空字符串方便查看
            row_data = [str(cell).strip() if cell is not None else "" for cell in row[:10]]
            # 只有整行不全为空才打印
            if any(row_data):
                print(f"Row {row_idx}: {row_data}")
except Exception as e:
    print(f"Error: {e}")
