import pandas as pd
from openpyxl import Workbook

# 创建一个简单的空模板，用于测试
wb = Workbook()
ws = wb.active
ws.title = "报价单"

# 设置一些占位表头
ws.cell(row=1, column=8, value="编号: [占位]")
ws.cell(row=2, column=1, value="客户：")
ws.cell(row=2, column=2, value="")
ws.cell(row=2, column=4, value="分析日期：")
ws.cell(row=3, column=1, value="结构：")
ws.cell(row=3, column=4, value="品名规格：")

# 制程表头
headers = ["序号", "制程", "详细规格", "", "", "", "单位用量", "单价", "材料金额"]
for col, h in enumerate(headers, 1):
    ws.cell(row=5, column=col, value=h)

wb.save("template/baojia_template.xlsx")
print("测试模板 template/baojia_template.xlsx 已生成")
