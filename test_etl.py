import sys
from openpyxl import load_workbook
import uuid
import re

file_path = "data/original_excels/171a94d36b5c480a8c3b972d6e73cea2.xlsx"
wb = load_workbook(file_path, data_only=True)

processed_count = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        first_cell = str(row[0]).strip() if row[0] else ""
        
        # 尝试去掉所有空格再匹配，解决 "成 本 分 析 表" 的问题
        clean_first = first_cell.replace(" ", "")
        if "成本分析表" in clean_first:
            print(f"\n--- Found anchor in sheet: {sheet_name} at row {row_idx} ---")
            try:
                # --- A. 提取基础信息 ---
                quotation_no = str(ws.cell(row=row_idx + 1, column=9).value or f"AUTO-{uuid.uuid4().hex[:6]}")
                print(f"quotation_no: {quotation_no}")
                
                customer = str(ws.cell(row=row_idx + 2, column=2).value or "")
                print(f"customer: {customer}")
                
                date_val = ws.cell(row=row_idx + 2, column=8).value
                print(f"date_val: {date_val}")
                
                structure = str(ws.cell(row=row_idx + 3, column=2).value or "")
                product_spec = str(ws.cell(row=row_idx + 3, column=8).value or "")
                print(f"product_spec: {product_spec}")
                
                # --- B. 正则提取业务标签 ---
                cross_section_match = re.search(r'(\d+(?:\.\d+)?mm2?)', product_spec, re.IGNORECASE)
                cross_section = cross_section_match.group(1) if cross_section_match else None
                
                material_codes = []
                materials_data = []
                mat_row = row_idx + 6
                
                while mat_row <= ws.max_row:
                    cell_a = str(ws.cell(row=mat_row, column=1).value or "").strip()
                    if "材料成本" in cell_a or not cell_a:
                        break
                        
                    process_name = str(ws.cell(row=mat_row, column=2).value or "")
                    spec_detail = str(ws.cell(row=mat_row, column=3).value or "")
                    
                    codes = re.findall(r'(EX\d+|D\d+[A-Z]?)', spec_detail)
                    material_codes.extend(codes)
                    
                    materials_data.append({
                        "unit_usage": ws.cell(row=mat_row, column=7).value,
                        "unit_price": ws.cell(row=mat_row, column=8).value,
                        "material_amount": ws.cell(row=mat_row, column=9).value,
                    })
                    mat_row += 1
                
                print(f"Found {len(materials_data)} materials")
                for mat in materials_data:
                    # Test float conversion
                    u1 = float(mat["unit_usage"]) if mat["unit_usage"] else 0.0
                    u2 = float(mat["unit_price"]) if mat["unit_price"] else 0.0
                    u3 = float(mat["material_amount"]) if mat["material_amount"] else 0.0
                print("Float conversion OK")
                processed_count += 1
                
            except Exception as e:
                print(f"Exception during parsing: {e}")

print(f"\nTotal processed: {processed_count}")
