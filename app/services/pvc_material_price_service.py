import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy import or_, text
from sqlalchemy.orm import Session

from app.models.pvc_material_price import PVCMaterialPrice, PVCMaterialPriceLog


def list_material_prices(db: Session, keyword: str = ""):
    keyword = (keyword or "").strip()
    rows = db.execute(text("""
        WITH bom_materials AS (
            SELECT
                PRD_NO,
                MIN(ZJNAME) AS ZJNAME,
                MIN(UT) AS BOM_UT,
                COUNT(*) AS used_count
            FROM dbo.PVC_BOM_Detail
            WHERE PRD_NO IS NOT NULL AND LTRIM(RTRIM(PRD_NO)) <> ''
            GROUP BY PRD_NO
        ),
        current_prices AS (
            SELECT *,
                   ROW_NUMBER() OVER (
                       PARTITION BY PRD_NO
                       ORDER BY
                           CASE WHEN MODIFYDATE IS NULL THEN 1 ELSE 0 END,
                           MODIFYDATE DESC,
                           CREATEDATE DESC,
                           id DESC
                   ) AS rn
            FROM dbo.PVC_MaterialPrice
        ),
        all_materials AS (
            SELECT PRD_NO FROM bom_materials
            UNION
            SELECT PRD_NO
            FROM dbo.PVC_MaterialPrice
            WHERE PRD_NO IS NOT NULL AND LTRIM(RTRIM(PRD_NO)) <> ''
        )
        SELECT
            p.id,
            a.PRD_NO,
            COALESCE(NULLIF(p.NAME, ''), b.ZJNAME, a.PRD_NO) AS NAME,
            COALESCE(NULLIF(p.UT, ''), b.BOM_UT) AS UT,
            p.UP,
            p.CREATEDATE,
            p.USR,
            p.MODIFYDATE,
            p.REM,
            p.HSYF,
            COALESCE(b.used_count, 0) AS used_count
        FROM all_materials a
        LEFT JOIN bom_materials b ON b.PRD_NO = a.PRD_NO
        LEFT JOIN current_prices p ON p.PRD_NO = a.PRD_NO AND p.rn = 1
        WHERE (:keyword = ''
            OR a.PRD_NO LIKE :like_keyword
            OR b.ZJNAME LIKE :like_keyword
            OR p.NAME LIKE :like_keyword)
        ORDER BY CASE WHEN p.UP IS NULL THEN 0 ELSE 1 END, a.PRD_NO
    """), {"keyword": keyword, "like_keyword": f"%{keyword}%"}).mappings().all()
    return [_serialize_current_material(row) for row in rows]


def create_material_price(db: Session, data: dict, operator: str, action: str = "CREATE", commit: bool = True):
    normalized = _normalize_payload(data)
    _ensure_not_duplicate(db, normalized)
    row = PVCMaterialPrice(**normalized, USR=operator, MODIFYDATE=datetime.now())
    db.add(row)
    db.flush()
    _write_log(db, row.id, row.PRD_NO, action, None, _snapshot(row), operator)
    if commit:
        db.commit()
    return row


def update_material_price(db: Session, row: PVCMaterialPrice, data: dict, operator: str):
    before = _snapshot(row)
    normalized = _normalize_payload(data, row)
    _ensure_not_duplicate(db, normalized, exclude_id=row.id)
    for key, value in normalized.items():
        setattr(row, key, value)
    row.USR = operator
    row.MODIFYDATE = datetime.now()
    _write_log(db, row.id, row.PRD_NO, "UPDATE", before, _snapshot(row), operator)
    db.commit()
    return row


def upsert_material_price(db: Session, data: dict, operator: str, commit: bool = True):
    normalized = _normalize_payload(data)
    row = (
        db.query(PVCMaterialPrice)
        .filter(PVCMaterialPrice.PRD_NO == normalized["PRD_NO"])
        .order_by(PVCMaterialPrice.id.desc())
        .first()
    )
    if row:
        for key, value in normalized.items():
            setattr(row, key, value)
        row.USR = operator
        row.MODIFYDATE = datetime.now()
        action = "updated"
        duplicate_rows = (
            db.query(PVCMaterialPrice)
            .filter(PVCMaterialPrice.PRD_NO == normalized["PRD_NO"], PVCMaterialPrice.id != row.id)
            .all()
        )
        for duplicate in duplicate_rows:
            db.delete(duplicate)
    else:
        row = PVCMaterialPrice(**normalized, USR=operator, MODIFYDATE=datetime.now())
        db.add(row)
        action = "created"
    if commit:
        db.commit()
    return row, action


def delete_material_price(db: Session, row: PVCMaterialPrice, operator: str):
    before = _snapshot(row)
    _write_log(db, row.id, row.PRD_NO, "DELETE", before, None, operator)
    db.delete(row)
    db.commit()


def list_material_price_logs(db: Session, prd_no: str = ""):
    query = db.query(PVCMaterialPriceLog)
    if prd_no:
        query = query.filter(PVCMaterialPriceLog.PRD_NO == prd_no.strip())
    return [
        serialize_log(row)
        for row in query.order_by(PVCMaterialPriceLog.operate_time.desc(), PVCMaterialPriceLog.id.desc()).limit(500).all()
    ]


def initialize_from_bom(db: Session, operator: str = "SYSTEM") -> dict:
    existing = {
        (row.PRD_NO, row.UT or "", _decimal_text(row.UP))
        for row in db.query(PVCMaterialPrice).all()
    }
    source_rows = db.execute(text("""
        SELECT PRD_NO, MIN(ZJNAME) AS NAME, UT, UP
        FROM dbo.PVC_BOM_Detail
        WHERE PRD_NO IS NOT NULL AND LTRIM(RTRIM(PRD_NO)) <> '' AND UP IS NOT NULL
        GROUP BY PRD_NO, UT, UP
        ORDER BY PRD_NO, UT, UP
    """)).mappings().all()
    missing_rows = db.execute(text("""
        SELECT DISTINCT PRD_NO, ZJNAME AS NAME, UT
        FROM dbo.PVC_BOM_Detail
        WHERE PRD_NO IS NOT NULL AND UP IS NULL
        ORDER BY PRD_NO
    """)).mappings().all()
    created = 0
    skipped = 0
    for source in source_rows:
        key = (source["PRD_NO"], source["UT"] or "", _decimal_text(source["UP"]))
        if key in existing:
            skipped += 1
            continue
        create_material_price(db, {
            "prd_no": source["PRD_NO"],
            "name": source["NAME"],
            "unit": source["UT"],
            "unit_price": source["UP"],
            "remark": "由历史 PVC BOM 明细初始化",
        }, operator, action="IMPORT", commit=False)
        existing.add(key)
        created += 1
    db.commit()
    return {
        "created": created,
        "skipped": skipped,
        "missing_price_items": [dict(row) for row in missing_rows],
    }


def import_material_prices_from_excel(db: Session, file_path: str, operator: str) -> dict:
    workbook = load_workbook(file_path, data_only=True)
    worksheet = workbook.active
    try:
        header_row_idx, columns = _find_import_columns(worksheet)
        created = 0
        updated = 0
        skipped = 0
        errors = []
        for row_idx in range(header_row_idx + 1, worksheet.max_row + 1):
            values = {key: worksheet.cell(row=row_idx, column=col_idx).value for key, col_idx in columns.items()}
            prd_no = str(values.get("prd_no") or "").strip()
            name = str(values.get("name") or "").strip()
            unit = str(values.get("unit") or "").strip()
            unit_price = values.get("unit_price")
            if not prd_no and not name and not unit and unit_price in (None, ""):
                continue
            if not prd_no or unit_price in (None, ""):
                skipped += 1
                errors.append(f"第 {row_idx} 行缺少材料代号或单价")
                continue
            if not name:
                name = _lookup_bom_material_name(db, prd_no)
            if not unit:
                unit = _lookup_bom_material_unit(db, prd_no)
            try:
                _, action = upsert_material_price(db, {
                    "prd_no": prd_no,
                    "name": name,
                    "unit": unit,
                    "unit_price": unit_price,
                    "remark": f"由 Excel 覆盖导入：{datetime.now():%Y-%m-%d %H:%M:%S}",
                }, operator, commit=False)
                if action == "created":
                    created += 1
                else:
                    updated += 1
            except ValueError as exc:
                skipped += 1
                errors.append(f"第 {row_idx} 行：{exc}")
        db.commit()
        return {"created": created, "updated": updated, "skipped": skipped, "errors": errors[:20]}
    finally:
        workbook.close()


def serialize_material_price(row: PVCMaterialPrice) -> dict:
    return {
        "id": row.id,
        "prd_no": row.PRD_NO,
        "name": row.NAME or "",
        "unit": row.UT or "",
        "unit_price": _decimal_text(row.UP),
        "effective_date": row.HSYF.date().isoformat() if row.HSYF else None,
        "remark": row.REM or "",
        "operator": row.USR or "",
        "create_time": row.CREATEDATE.isoformat() if row.CREATEDATE else None,
        "update_time": row.MODIFYDATE.isoformat() if row.MODIFYDATE else None,
    }


def _serialize_current_material(row) -> dict:
    return {
        "id": row["id"],
        "prd_no": row["PRD_NO"] or "",
        "name": row["NAME"] or "",
        "unit": row["UT"] or "",
        "unit_price": _decimal_text(row["UP"]),
        "has_price": row["UP"] is not None,
        "used_count": int(row["used_count"] or 0),
        "remark": row["REM"] or "",
        "operator": row["USR"] or "",
        "create_time": row["CREATEDATE"].isoformat() if row["CREATEDATE"] else None,
        "update_time": row["MODIFYDATE"].isoformat() if row["MODIFYDATE"] else None,
    }


def serialize_log(row: PVCMaterialPriceLog) -> dict:
    return {
        "id": row.id,
        "material_price_id": row.material_price_id,
        "prd_no": row.PRD_NO,
        "action": row.action,
        "before_data": json.loads(row.before_data) if row.before_data else None,
        "after_data": json.loads(row.after_data) if row.after_data else None,
        "operator": row.operator,
        "operate_time": row.operate_time.isoformat() if row.operate_time else None,
    }


def _normalize_payload(data: dict, existing: PVCMaterialPrice | None = None) -> dict:
    prd_no = str(data.get("prd_no", existing.PRD_NO if existing else "") or "").strip()
    name = str(data.get("name", existing.NAME if existing else "") or "").strip()
    unit = str(data.get("unit", existing.UT if existing else "") or "").strip().upper()
    price_raw = data.get("unit_price", existing.UP if existing else None)
    if not prd_no or not name or not unit or price_raw in (None, ""):
        raise ValueError("材料编号、材料名称、材料单位和未税单价均为必填项")
    try:
        unit_price = Decimal(str(price_raw))
    except InvalidOperation as exc:
        raise ValueError("未税单价格式不正确") from exc
    if unit_price < 0:
        raise ValueError("未税单价不能小于 0")
    effective_raw = data.get("effective_date", existing.HSYF if existing else None)
    effective_date = _parse_date(effective_raw)
    return {
        "PRD_NO": prd_no,
        "NAME": name,
        "UT": unit,
        "UP": unit_price,
        "HSYF": effective_date,
        "REM": str(data.get("remark", existing.REM if existing else "") or "").strip(),
    }


def _find_import_columns(worksheet):
    aliases = {
        "prd_no": {"材料代号", "材料编号", "材料编码", "料号", "物料编号", "物料编码", "PRD_NO"},
        "name": {"材料名称", "物料名称", "名称", "NAME"},
        "unit": {"材料单位", "单位", "物料单位", "UT"},
        "unit_price": {"单价", "未税单价", "材料单价", "最新单价", "UP"},
    }
    normalized_aliases = {
        key: {_normalize_header(value) for value in values}
        for key, values in aliases.items()
    }
    for row_idx in range(1, min(worksheet.max_row, 20) + 1):
        found = {}
        for col_idx in range(1, worksheet.max_column + 1):
            header = _normalize_header(worksheet.cell(row=row_idx, column=col_idx).value)
            if not header:
                continue
            for key, values in normalized_aliases.items():
                if header in values and key not in found:
                    found[key] = col_idx
        if {"prd_no", "unit_price"}.issubset(found):
            return row_idx, found
    raise ValueError("未识别到 Excel 表头，至少需要材料代号和单价列")


def _normalize_header(value) -> str:
    return str(value or "").strip().replace(" ", "").replace("\n", "").upper()


def _lookup_bom_material_name(db: Session, prd_no: str) -> str:
    return db.execute(text("""
        SELECT TOP 1 ZJNAME FROM dbo.PVC_BOM_Detail
        WHERE PRD_NO = :prd_no AND ZJNAME IS NOT NULL
        ORDER BY id
    """), {"prd_no": prd_no}).scalar() or prd_no


def _lookup_bom_material_unit(db: Session, prd_no: str) -> str:
    return db.execute(text("""
        SELECT TOP 1 UT FROM dbo.PVC_BOM_Detail
        WHERE PRD_NO = :prd_no AND UT IS NOT NULL
        ORDER BY id
    """), {"prd_no": prd_no}).scalar() or "KG"


def _parse_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    try:
        return datetime.fromisoformat(str(value))
    except ValueError as exc:
        raise ValueError("生效日期格式不正确") from exc


def _ensure_not_duplicate(db: Session, data: dict, exclude_id: int | None = None):
    query = db.query(PVCMaterialPrice).filter(
        PVCMaterialPrice.PRD_NO == data["PRD_NO"],
        PVCMaterialPrice.UT == data["UT"],
        PVCMaterialPrice.UP == data["UP"],
    )
    if data["HSYF"] is None:
        query = query.filter(PVCMaterialPrice.HSYF.is_(None))
    else:
        query = query.filter(PVCMaterialPrice.HSYF == data["HSYF"])
    if exclude_id:
        query = query.filter(PVCMaterialPrice.id != exclude_id)
    if query.first():
        raise ValueError("相同料号、单位、单价和生效日期的价格版本已存在")


def _write_log(db: Session, row_id: int, prd_no: str, action: str, before: dict | None, after: dict | None, operator: str):
    db.add(PVCMaterialPriceLog(
        material_price_id=row_id,
        PRD_NO=prd_no,
        action=action,
        before_data=json.dumps(before, ensure_ascii=False) if before else None,
        after_data=json.dumps(after, ensure_ascii=False) if after else None,
        operator=operator,
    ))


def _snapshot(row: PVCMaterialPrice) -> dict:
    return serialize_material_price(row)


def _decimal_text(value) -> str | None:
    if value is None:
        return None
    return f"{Decimal(value):f}".rstrip("0").rstrip(".") or "0"
