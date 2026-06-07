import os
import re
import uuid
import time
import json
import shutil
import asyncio
import logging
import hashlib
from dataclasses import dataclass
from datetime import datetime
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session
from fastapi import UploadFile
from openai import AsyncOpenAI

from app.models.calc_param import QuotationCalcParam
from app.models.calculation_trace import QuotationCalculationTrace
from app.models.quotation import QuotationBpmInstance, QuotationMain, QuotationMaterial, QuotationProcessFee
from app.models.user import User
from app.core.config import settings
from app.services.bpm_instance_service import (
    REVIEW_PENDING,
    REVIEW_QUOTED,
    ensure_bpm_instance,
    get_existing_quotation,
    normalize_bpm_no,
)
from app.services.quotation_summary_service import apply_quotation_summaries

logger = logging.getLogger(__name__)

async_client = AsyncOpenAI(
    api_key=settings.DEEPSEEK_API_KEY,
    base_url=settings.DEEPSEEK_BASE_URL
)

UPLOAD_DIR = "data/original_excels"
os.makedirs(UPLOAD_DIR, exist_ok=True)

_BRAIDING_RE = re.compile(r'(\d+(?:\.\d+)?)\s*%?\s*编')

EXTRACTION_PROMPT = """你是一个制造业电缆报价单数据提取专家。
请从下面的【表头区数据】和【核心表格数据】中，提取指定的字段，并严格以 JSON 格式输出。

注意规则：
1. 绝对不能修改任何数值，保持原始浮点数精度！如果某个数值为空或解析不到，请使用 0.0 或 null。
2. 从详细规格中提取出所有胶料料号（常见前缀: EX, EE, D, V, C, PA），存入 material_codes 数组。
3. 将百分比转换为小数（如 2% 转换为 0.02）。如果本身是小数则保持不变。
4. 表头区"结构"列的右侧紧跟"编织率(%)"列，提取编织率百分比值，转换为小数（如 95 → 0.95、85% → 0.85），存入 braiding_rate。如果该列为空，再尝试从品名规格中提取（如 "95%编织"）。
5. 新模板表头区包含 "客户名称"、"收货地（市）"、"包装方式-米数"、"备注" 等字段：客户名称存入 customer，收货地（市）存入 address，包装方式-米数存入 package_method，备注存入 remark。
6. 旧模板中如果仍出现 "客户"、"地址"，按 customer/address 兼容提取。
7. 核心表格中每一行制程的规格列之后紧跟"物料编码"列，将物料编码提取到该行 material 的 material_code 字段。物料编码可能是空值或"新开发"字样，均如实提取。
8. 【非常重要】务必准确提取 "quotation_no" (编号)！在表头区，编号可能出现在 "编号:"、"编号：" 之后，或者孤立地出现在某一列（如 FHLR2GCB2G-50-003）。请仔细扫描【表头区数据】的所有列，不要遗漏。

【表头区数据】
{header_text}

【核心表格数据】
{material_text}

请严格输出如下 JSON 格式（【重要】只输出纯 JSON，禁止用 ``` 包裹，禁止添加任何解释文字，响应第一个字符必须是 {{）：
{{
    "quotation_no": "编号",
    "customer": "客户",
    "address": "收货地（市）/地址",
    "package_method": "包装方式-米数",
    "date_val": "分析日期 (YYYY-MM-DD)",
    "structure": "结构",
    "product_spec": "品名规格",
    "braiding_rate": 浮点数 (编织率，小数形式，如 0.95),
    "remark": "备注",
    "material_codes": ["料号1", "料号2"],
    "materials": [
        {{
            "process_name": "制程名称",
            "spec_detail": "详细规格",
            "material_code": "物料编码（可能为空或"新开发"）",
            "unit_usage": 浮点数 (单位用量),
            "unit_price": 浮点数 (单价),
            "material_amount": 浮点数 (材料金额)
        }}
    ],
    "processes": [
        {{
            "process_name": "制程名称",
            "std_hours": 浮点数 (标准工时),
            "loss_hours": 浮点数 (损耗时间),
            "fixed_rate": 浮点数 (固定费用率),
            "fixed_cost": 浮点数 (固定费用),
            "startup_loss_wire": 浮点数 (开机损耗废线),
            "total_waste_glue": 浮点数 (每个制程总废胶),
            "amount": 浮点数 (金额),
            "subtotal_cost": 浮点数 (费用成本小计)
        }}
    ],
    "cost_summary": {{
        "total_material_cost_rmb_m": 浮点数 (材料成本 RMB/M),
        "total_material_cost_kg": 浮点数 (材料成本 Kg),
        "total_material_amount": 浮点数 (材料成本总金额),
        "total_process_cost": 浮点数 (费用总计),
        "ul_label_fee": 浮点数,
        "transport_fee": 浮点数,
        "package_fee": 浮点数,
        "scrap_rate": 浮点数,
        "startup_times": 整数 (订单开机次数),
        "delivery_fee": 浮点数,
        "customs_fee": 浮点数,
        "order_meters": 整数,
        "net_profit_rate": 浮点数,
        "vat_rate": 浮点数,
        "business_fee_rate": 浮点数,
        "monthly_interest_rate": 浮点数,
        "corp_tax_rate": 浮点数,
        "irradiation_core_count": 浮点数 (照射芯数),
        "irradiation_core_fee": 浮点数 (照射费用 RMB/M),
        "other_fee": 浮点数 (其他费用),
        "cost_rmb_m": 浮点数 (底部汇总的 成本 RMB/M),
        "price_with_profit": 浮点数 (取利售价),
        "price_without_profit": 浮点数 (不取利售价),
        "final_price": 浮点数 (最终售价)
    }}
}}"""

_mat_code_pattern = re.compile(r'((?:EX|EE|PA|D|V|C)\d+[A-Z]*)', re.IGNORECASE)


def _admin_creator_names(db: Session, tenant_id: str) -> list[str]:
    rows = (
        db.query(User.username, User.display_name)
        .filter(User.is_admin == True)
        .all()
    )
    names = set()
    for username, display_name in rows:
        if username:
            names.add(username)
        if display_name:
            names.add(display_name)
    return list(names)


@dataclass
class QuotationBlock:
    """Excel 中一个报价单的文本块"""
    sheet_name: str
    row_idx: int
    header_text: str
    table_text: str
    preview: str  # 编号预览，用于进度展示


def _parse_braiding_rate(product_spec: str) -> float:
    """从品名规格中提取编织率，返回小数形式（95% → 0.95）"""
    m = _BRAIDING_RE.search(product_spec or "")
    if not m:
        return 0.0
    val = float(m.group(1))
    return val / 100 if val > 1 else val


def _normalize_braiding_rate(val) -> float:
    """统一编织率为 0-1 小数：如果 >1 则视为百分比除以 100"""
    v = _safe_numeric(val, scale=4)
    if v > 1:
        v = round(v / 100, 4)
    return v


def _safe_numeric(val, scale: int = 4) -> float:
    if val is None:
        return 0.0
    try:
        return round(float(val), scale)
    except (ValueError, TypeError):
        return 0.0


def _first_json_value(value):
    if isinstance(value, list):
        for item in value:
            if item is not None:
                return item
        return None
    if isinstance(value, dict):
        for item in value.values():
            if item is not None:
                return item
        return None
    return value


def _normalize_extraction_shape(data: dict) -> dict:
    """Normalize LLM JSON into the shape the DB writer expects."""
    if not isinstance(data, dict):
        raise ValueError(f"Expected extraction result dict, got {type(data).__name__}")

    scalar_fields = [
        "quotation_no",
        "customer",
        "address",
        "package_method",
        "date_val",
        "structure",
        "product_spec",
        "braiding_rate",
        "remark",
    ]
    for field in scalar_fields:
        if field in data:
            data[field] = _first_json_value(data[field])

    raw_codes = data.get("material_codes", [])
    if isinstance(raw_codes, str):
        data["material_codes"] = [raw_codes.strip()] if raw_codes.strip() else []
    elif isinstance(raw_codes, list):
        data["material_codes"] = [str(item).strip() for item in raw_codes if str(item or "").strip()]
    else:
        data["material_codes"] = []

    cost_summary = data.get("cost_summary", {})
    if isinstance(cost_summary, list):
        cost_summary = next((item for item in cost_summary if isinstance(item, dict)), {})
    data["cost_summary"] = cost_summary if isinstance(cost_summary, dict) else {}

    for list_field in ("materials", "processes"):
        rows = data.get(list_field, [])
        if isinstance(rows, dict):
            rows = [rows]
        elif not isinstance(rows, list):
            rows = []
        normalized_rows = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            normalized_rows.append({
                key: _first_json_value(value)
                for key, value in row.items()
            })
        data[list_field] = normalized_rows

    return data


def _parse_excel_numeric(val):
    """Parse raw Excel numeric values, including text percentages like ``17%``."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    text = str(val).strip()
    if not text:
        return None
    text = text.replace(",", "").replace("，", "")
    try:
        if text.endswith("%"):
            return float(text[:-1].strip()) / 100
        return float(text)
    except (ValueError, TypeError):
        return None


def _compact_cell_text(val) -> str:
    if val is None:
        return ""
    return re.sub(r"\s+", "", str(val)).strip()


def _block_content_hash(block: QuotationBlock | None) -> str:
    if not block:
        return ""
    text_value = "\n".join([block.header_text or "", block.table_text or ""])
    text_value = re.sub(r"\d{4}[/-]\d{1,2}[/-]\d{1,2}", "", text_value)
    text_value = re.sub(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", "", text_value)
    text_value = re.sub(r"\s+", "", text_value).upper()
    return hashlib.sha256(text_value.encode("utf-8", errors="ignore")).hexdigest()


def _parse_date_value(value):
    if not value:
        return None
    if hasattr(value, "date"):
        try:
            return value.date()
        except Exception:
            pass
    text_value = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(text_value, fmt).date()
        except Exception:
            continue
    return None


def _quick_header_from_excel(saved_path: str, block: QuotationBlock) -> dict:
    data = _patch_header_fields_from_excel({}, saved_path, block)
    return data if isinstance(data, dict) else {}


def _patch_header_fields_from_excel(data: dict, file_path: str, block: QuotationBlock | None) -> dict:
    """Patch header metadata from the uploaded workbook using template labels.

    This keeps new-template fields deterministic and avoids depending only on LLM
    interpretation for short labels such as "收货地（市）" and "包装方式-米数".
    """
    if data is None or not file_path or not block:
        return data

    ext = os.path.splitext(file_path)[1].lower()
    is_xls = ext == ".xls"
    wb = None
    try:
        if is_xls:
            import xlrd
            wb = xlrd.open_workbook(file_path)
            ws = wb.sheet_by_name(block.sheet_name)
            max_row = ws.nrows
            max_col = ws.ncols

            def cell_value(row: int, col: int):
                if row < 1 or col < 1 or row > max_row or col > max_col:
                    return None
                return ws.cell_value(row - 1, col - 1)
        else:
            wb = load_workbook(file_path, data_only=True)
            ws = wb[block.sheet_name]
            max_row = ws.max_row
            max_col = ws.max_column

            def cell_value(row: int, col: int):
                if row < 1 or col < 1 or row > max_row or col > max_col:
                    return None
                return ws.cell(row=row, column=col).value

        label_map = {
            "quotation_no": ("编号",),
            "customer": ("客户名称", "客户"),
            "address": ("收货地（市）", "收货地(市)", "收货地", "地址"),
            "package_method": ("包装方式-米数", "包装方式"),
            "date_val": ("分析日期",),
            "structure": ("结构",),
            "braiding_rate": ("编织率",),
            "product_spec": ("品名规格",),
            "remark": ("备注",),
        }

        patched: dict[str, str] = {}
        start_row = max(1, block.row_idx)
        end_row = min(max_row, block.row_idx + 24)
        for row in range(start_row, end_row + 1):
            for col in range(1, max_col + 1):
                label_text = _compact_cell_text(cell_value(row, col)).rstrip(":：")
                if not label_text:
                    continue
                for field, labels in label_map.items():
                    if field in patched:
                        continue
                    if not any(label in label_text for label in labels):
                        continue
                    value = _next_non_empty_cell_text(cell_value, row, col, max_col)
                    if value:
                        patched[field] = value

        if patched:
            for field, value in patched.items():
                data[field] = value
            logger.info(
                "[%s] Excel structure patched header fields: %s",
                data.get("quotation_no") or block.preview[:50],
                ", ".join(sorted(patched.keys())),
            )
    except Exception as e:
        logger.warning(
            "[%s] Excel header patch skipped: %s",
            data.get("quotation_no") if data else block.preview[:50],
            e,
        )
    finally:
        if wb is not None and not is_xls:
            wb.close()

    return data


def _next_non_empty_cell_text(cell_value, row: int, label_col: int, max_col: int) -> str:
    for col in range(label_col + 1, min(max_col, label_col + 4) + 1):
        text_value = str(cell_value(row, col) or "").strip()
        if text_value:
            return text_value
    return ""


def _patch_cost_summary_from_excel(data: dict, file_path: str, block: QuotationBlock | None) -> dict:
    """Use fixed Excel layout to correct bottom fee fields after LLM extraction.

    The template has two visible "other fee" labels: a large merged area label and
    the actual field header. Matching the header row shape keeps the stored value
    aligned with the real editable field.
    """
    if not data or not file_path or not block:
        return data

    ext = os.path.splitext(file_path)[1].lower()
    is_xls = ext == ".xls"
    wb = None
    try:
        if is_xls:
            import xlrd
            wb = xlrd.open_workbook(file_path)
            ws = wb.sheet_by_name(block.sheet_name)
            max_row = ws.nrows
            max_col = ws.ncols

            def cell_value(row: int, col: int):
                if row < 1 or col < 1 or row > max_row or col > max_col:
                    return None
                return ws.cell_value(row - 1, col - 1)
        else:
            wb = load_workbook(file_path, data_only=True)
            ws = wb[block.sheet_name]
            max_row = ws.max_row
            max_col = ws.max_column

            def cell_value(row: int, col: int):
                if row < 1 or col < 1 or row > max_row or col > max_col:
                    return None
                return ws.cell(row=row, column=col).value

        cost_summary = data.setdefault("cost_summary", {})

        header_groups = [
            {
                "anchors": ("UL标签费", "运输费", "包装费"),
                "fields": {
                    "ul_label_fee": "UL标签费",
                    "transport_fee": "运输费",
                    "package_fee": "包装费",
                    "scrap_rate": "废品损耗",
                    "startup_times": "订单开机次数",
                    "total_process_cost": "费用总计",
                },
            },
            {
                "anchors": ("其他费用", "净利率", "报关费", "订单米数"),
                "fields": {
                    "other_fee": "其他费用",
                    "net_profit_rate": "净利率",
                    "customs_fee": "报关费",
                    "vat_rate": "增值税率",
                    "order_meters": "订单米数",
                },
            },
            {
                "anchors": ("照射芯数", "照射费用", "营业费用率"),
                "fields": {
                    "irradiation_core_count": "照射芯数",
                    "irradiation_core_fee": "照射费用",
                    "business_fee_rate": "营业费用率",
                    "monthly_interest_rate": "月结利息",
                    "corp_tax_rate": "企税税率",
                },
            },
        ]

        patched: dict[str, float] = {}
        start_row = max(1, block.row_idx)
        end_row = min(max_row, block.row_idx + 80)
        for row in range(start_row, end_row + 1):
            cells = [_compact_cell_text(cell_value(row, col)) for col in range(1, max_col + 1)]
            row_text = "|".join(cells)
            for group in header_groups:
                if not all(anchor in row_text for anchor in group["anchors"]):
                    continue
                for key, header in group["fields"].items():
                    for idx, cell_text in enumerate(cells, start=1):
                        if header in cell_text:
                            parsed = _parse_excel_numeric(cell_value(row + 1, idx))
                            if parsed is not None:
                                cost_summary[key] = parsed
                                patched[key] = parsed
                            break

        if patched:
            logger.info(
                "[%s] Excel structure patched cost_summary fields: %s",
                data.get("quotation_no") or block.preview[:50],
                ", ".join(sorted(patched.keys())),
            )
    except Exception as e:
        logger.warning(
            "[%s] Excel structure patch skipped: %s",
            data.get("quotation_no") if data else block.preview[:50],
            e,
        )
    finally:
        if wb is not None and not is_xls:
            wb.close()

    return data


def _row_values_xlsx(ws, row_idx: int) -> list[str]:
    """openpyxl: 取一行所有单元格的字符串值（1-indexed）"""
    return [str(c.value).strip() if c.value is not None else "" for c in ws[row_idx]]


def _row_values_xls(ws, row_idx: int) -> list[str]:
    """xlrd: 取一行所有单元格的字符串值（0-indexed）"""
    return [str(v).strip() for v in ws.row_values(row_idx)]


def scan_quotations(file_path: str) -> list[QuotationBlock]:
    """阶段1：扫描 Excel(.xlsx/.xls)，定位所有 成本分析表 锚点并提取文本块"""
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(file_path)
        sheet_names = wb.sheet_names()
        row_values = lambda ws, r: _row_values_xls(ws, r - 1)  # xlrd 0-indexed
        max_row = lambda ws: ws.nrows
        is_xls = True
    else:
        wb = load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        row_values = _row_values_xlsx
        max_row = lambda ws: ws.max_row
        is_xls = False

    blocks = []

    for sheet_name in sheet_names:
        ws = wb.sheet_by_name(sheet_name) if is_xls else wb[sheet_name]

        for row_idx in range(1, max_row(ws) + 1):
            vals = row_values(ws, row_idx)
            first_cell = vals[0] if vals else ""
            clean_first = first_cell.replace(" ", "")

            if "成本分析表" in clean_first:
                # 提取表头区（下 4 行）
                header_rows = []
                for i in range(1, 5):
                    header_rows.append(" | ".join(row_values(ws, row_idx + i)))
                header_text = "\n".join(header_rows)

                # 提取核心表格数据区
                table_rows = []
                mat_row = row_idx + 6
                empty_count = 0
                while mat_row <= max_row(ws):
                    row_vals = row_values(ws, mat_row)
                    cell_a = row_vals[0].replace(" ", "") if row_vals else ""

                    if "成本分析表" in cell_a:
                        break

                    if not any(row_vals):
                        empty_count += 1
                        if empty_count > 3:
                            break
                    else:
                        empty_count = 0

                    table_rows.append(" | ".join(row_vals))
                    mat_row += 1
                    if mat_row - row_idx > 80:
                        break
                table_text = "\n".join(table_rows)

                preview = header_rows[0][:60] if header_rows else f"Row {row_idx}"

                blocks.append(QuotationBlock(
                    sheet_name=sheet_name,
                    row_idx=row_idx,
                    header_text=header_text,
                    table_text=table_text,
                    preview=preview,
                ))

    if not is_xls:
        wb.close()
    return blocks


async def _extract_one_async(block: QuotationBlock, semaphore: asyncio.Semaphore) -> dict:
    """对单个报价单文本块调用 LLM 提取，受 semaphore 限制并发数。空内容时自动重试一次。"""
    import re as _re
    prompt = EXTRACTION_PROMPT.format(
        header_text=block.header_text,
        material_text=block.table_text,
    )

    def _parse_json(raw: str) -> dict:
        """尝试解析 LLM 返回内容为 JSON，支持去除 markdown 代码块包裹。确保始终返回 dict。"""
        if not raw or not raw.strip():
            raise ValueError("Empty or whitespace-only response")

        def _load(text: str):
            """json.loads + 容错：如果是数组则取第一个元素"""
            obj = json.loads(text)
            if isinstance(obj, list):
                if len(obj) > 0 and isinstance(obj[0], dict):
                    return obj[0]
                raise ValueError(f"Expected JSON object, got list with {len(obj)} non-dict items")
            if not isinstance(obj, dict):
                raise ValueError(f"Expected JSON object, got {type(obj).__name__}")
            return obj

        # 尝试直接解析
        try:
            return _load(raw)
        except (json.JSONDecodeError, ValueError):
            pass
        # 尝试提取 markdown 代码块中的 JSON
        m = _re.search(r'```(?:json)?\s*\n?(.*?)\n?```', raw, _re.DOTALL)
        if m:
            return _load(m.group(1).strip())
        # 尝试找第一个 { 到最后一个 }
        start = raw.find('{')
        end = raw.rfind('}')
        if start != -1 and end > start:
            return _load(raw[start:end + 1])
        raise

    async with semaphore:
        last_error = None
        for attempt in range(2):
            try:
                response = await async_client.chat.completions.create(
                    model=settings.DEEPSEEK_MODEL,
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.0,
                    max_tokens=4096,
                )
                choice = response.choices[0]
                result_str = choice.message.content
                if not result_str or not result_str.strip():
                    detail = {
                        "finish_reason": choice.finish_reason,
                        "content": result_str,
                        "refusal": getattr(choice.message, "refusal", None),
                        "reasoning_content": getattr(choice.message, "reasoning_content", None),
                        "usage": str(response.usage) if response.usage else None,
                        "prompt_len": len(prompt),
                        "attempt": attempt + 1,
                    }
                    logger.warning(
                        f"LLM returned empty content (attempt {attempt + 1}/2)! "
                        f"Detail: {json.dumps(detail, ensure_ascii=False)}"
                    )
                    if attempt == 0:
                        await asyncio.sleep(2)
                        continue
                    raise ValueError(f"LLM returned empty content after 2 attempts")

                return _normalize_extraction_shape(_parse_json(result_str))

            except json.JSONDecodeError:
                # 非空但格式不对：记录实际内容便于排查，不重试
                logger.error(
                    f"LLM returned non-JSON content (len={len(result_str)}): "
                    f"{str(result_str)[:500]}"
                )
                raise
            except ValueError:
                raise  # 空内容异常直接抛出
            except Exception as e:
                last_error = e
                logger.error(f"LLM call failed (attempt {attempt + 1}/2): {e}")
                if attempt == 0:
                    await asyncio.sleep(2)

        if last_error:
            raise last_error
        raise ValueError(f"LLM returned empty content after 2 attempts, prompt_len={len(prompt)}")


def _enrich_extracted(data: dict, product_spec: str) -> None:
    """正则兜底补充：material_codes 和 braiding_rate"""
    # 编织率兜底
    if _safe_numeric(data.get("braiding_rate"), scale=4) == 0.0:
        data["braiding_rate"] = _parse_braiding_rate(product_spec)

    # 材料料号兜底：从规格中扫描
    existing_codes = [c.upper() for c in data.get("material_codes", [])]
    for mat in data.get("materials", []):
        spec = mat.get("spec_detail", "")
        if spec:
            for code in _mat_code_pattern.findall(str(spec)):
                if code.upper() not in existing_codes:
                    existing_codes.append(code.upper())
                    data.setdefault("material_codes", []).append(code.upper())

        # 物料编码兜底：如果 LLM 没提取到 material_code，从规格中正则提取
        if not mat.get("material_code"):
            found = _mat_code_pattern.findall(str(mat.get("spec_detail", "")))
            if found:
                mat["material_code"] = found[0]


def _quotation_exists(db: Session, quotation_code: str, tenant_id: str) -> bool:
    """检查同租户下成本分析号是否已存在"""
    return db.query(QuotationMain).filter(
        QuotationMain.quotation_code == quotation_code,
        QuotationMain.tenant_id == tenant_id,
    ).first() is not None


def _delete_quotation_fk_children(db: Session, quotation_id: int) -> None:
    """Delete all direct children that reference quotation_main before deleting main."""
    rows = db.execute(text("""
        SELECT
            OBJECT_SCHEMA_NAME(fkc.parent_object_id) AS child_schema,
            OBJECT_NAME(fkc.parent_object_id) AS child_table,
            pc.name AS child_column
        FROM sys.foreign_key_columns fkc
        JOIN sys.columns pc
          ON pc.object_id = fkc.parent_object_id
         AND pc.column_id = fkc.parent_column_id
        WHERE OBJECT_NAME(fkc.referenced_object_id) = 'quotation_main'
        ORDER BY OBJECT_NAME(fkc.parent_object_id)
    """)).mappings().all()
    for row in rows:
        schema = str(row["child_schema"]).replace("]", "]]")
        table = str(row["child_table"]).replace("]", "]]")
        column = str(row["child_column"]).replace("]", "]]")
        db.execute(
            text(f"DELETE FROM [{schema}].[{table}] WHERE [{column}] = :quotation_id"),
            {"quotation_id": quotation_id},
        )


def _write_one_quotation(
    db: Session,
    data: dict,
    tenant_id: str,
    username: str,
    saved_path: str,
    display_name: str = "",
    bpm_no: str = "",
    content_hash: str = "",
) -> str:
    """将单条 LLM 提取结果写入数据库，返回 quotation_code。
    如果成本分析号已存在则跳过，返回空字符串。"""
    creator_name = display_name or username
    quotation_code = data.get("quotation_no") or f"AUTO-{uuid.uuid4().hex[:6]}"
    product_spec = data.get("product_spec") or ""

    # 唯一性校验：同租户下已存在则跳过
    existing = get_existing_quotation(db, tenant_id, quotation_code)
    if existing:
        ensure_bpm_instance(
            db,
            existing,
            bpm_no,
            _parse_date_value(data.get("date_val")) or existing.analysis_date,
            creator_name,
            saved_path,
        )
        if content_hash and not getattr(existing, "content_hash", None):
            existing.content_hash = content_hash
        existing.updater = creator_name
        existing.update_time = datetime.now()
        logger.info(f"[{quotation_code}] 成本分析号已存在，跳过")
        return ""

    _enrich_extracted(data, product_spec)

    cost_data = data.get("cost_summary", {})
    braiding_rate = _normalize_braiding_rate(data.get("braiding_rate"))

    # 日期处理
    date_val = data.get("date_val")
    try:
        parsed_date = datetime.strptime(date_val, "%Y-%m-%d").date() if date_val else datetime.now().date()
    except Exception:
        parsed_date = datetime.now().date()

    # extracted_tags
    material_codes = list(data.get("material_codes", []))
    extracted_tags = json.dumps({
        "material_codes": material_codes,
        "customer_keyword": (data.get("customer") or "").split()[0] if data.get("customer") else None,
        "package_method": data.get("package_method") or "",
    }, ensure_ascii=False)

    new_main = QuotationMain(
        tenant_id=tenant_id,
        quotation_code=quotation_code,
        bpm_no=normalize_bpm_no(bpm_no),
        customer_name=data.get("customer") or "",
        customer_address=data.get("address") or "",
        package_method=data.get("package_method") or "",
        analysis_date=parsed_date,
        structure=data.get("structure") or "",
        product_spec=product_spec,
        remark=data.get("remark") or "",
        original_file_path=saved_path,
        content_hash=content_hash or "",
        extracted_tags=extracted_tags,
        unit_usage_sum=None,
        material_amount_sum=_safe_numeric(cost_data.get("total_material_amount")),
        material_cost=_safe_numeric(cost_data.get("total_material_cost_rmb_m")),
        total_fee=_safe_numeric(cost_data.get("total_process_cost")),
        ul_label_fee=_safe_numeric(cost_data.get("ul_label_fee")),
        transport_fee=_safe_numeric(cost_data.get("transport_fee")),
        packing_fee=_safe_numeric(cost_data.get("package_fee")),
        waste_loss_rate=_safe_numeric(cost_data.get("scrap_rate")),
        order_startup_times=_safe_numeric(cost_data.get("startup_times"), scale=0),
        other_fee=_safe_numeric(cost_data.get("other_fee")),
        delivery_fee=_safe_numeric(cost_data.get("delivery_fee")),
        irradiation_core_count=_safe_numeric(cost_data.get("irradiation_core_count")),
        irradiation_core_fee=_safe_numeric(cost_data.get("irradiation_core_fee")),
        customs_fee=_safe_numeric(cost_data.get("customs_fee")),
        order_meterage=_safe_numeric(cost_data.get("order_meters")),
        net_profit_rate=_safe_numeric(cost_data.get("net_profit_rate")),
        vat_rate=_safe_numeric(cost_data.get("vat_rate")),
        operating_expense_rate=_safe_numeric(cost_data.get("business_fee_rate")),
        monthly_interest=_safe_numeric(cost_data.get("monthly_interest_rate")),
        corporate_tax_rate=_safe_numeric(cost_data.get("corp_tax_rate")),
        cost=_safe_numeric(cost_data.get("cost_rmb_m")),
        profit_selling_price=_safe_numeric(cost_data.get("price_with_profit")),
        non_profit_price=_safe_numeric(cost_data.get("price_without_profit")),
        final_selling_price=_safe_numeric(cost_data.get("final_price")),
        braiding_rate=braiding_rate,
        creator=creator_name,
        deleted=False,
    )
    db.add(new_main)
    db.flush()

    material_rows = []
    process_rows = []

    for idx, mat in enumerate(data.get("materials", [])):
        raw_code = mat.get("material_code")
        material_code = str(raw_code).strip() if raw_code and str(raw_code).strip().lower() != "null" else ""
        material_row = QuotationMaterial(
            tenant_id=tenant_id,
            quotation_main_id=new_main.id,
            seq_no=idx + 1,
            process_name=mat.get("process_name") or "",
            spec_detail=mat.get("spec_detail") or "",
            unit_usage=_safe_numeric(mat.get("unit_usage")),
            unit_price=_safe_numeric(mat.get("unit_price")),
            material_amount=_safe_numeric(mat.get("material_amount")),
            process_code=material_code,
            creator=creator_name,
            deleted=False,
        )
        material_rows.append(material_row)
        db.add(material_row)

    for proc in data.get("processes", []):
        process_row = QuotationProcessFee(
            tenant_id=tenant_id,
            quotation_main_id=new_main.id,
            process_name=proc.get("process_name") or "",
            std_hours=_safe_numeric(proc.get("std_hours")),
            loss_hours=_safe_numeric(proc.get("loss_hours")),
            fixed_rate=_safe_numeric(proc.get("fixed_rate")),
            fixed_fee=_safe_numeric(proc.get("fixed_cost")),
            startup_loss_wire=_safe_numeric(proc.get("startup_loss_wire")),
            total_waste_glue=_safe_numeric(proc.get("total_waste_glue")),
            amount=_safe_numeric(proc.get("amount")),
            subtotal_fee=_safe_numeric(proc.get("subtotal_cost")),
            creator=creator_name,
            deleted=False,
        )
        process_rows.append(process_row)
        db.add(process_row)

    apply_quotation_summaries(new_main, material_rows, process_rows)
    ensure_bpm_instance(db, new_main, bpm_no, parsed_date, creator_name, saved_path)

    return quotation_code


async def process_excel_streaming(
    blocks: list[QuotationBlock],
    saved_path: str,
    db: Session,
    tenant_id: str,
    username: str,
    display_name: str = "",
    bpm_no: str = "",
):
    """阶段2+3：并行调用 LLM 提取，实时产出 SSE 进度事件，最后写库（含去重）"""
    total = len(blocks)

    if total == 0:
        yield {"event": "complete", "processed": 0, "errors": 0}
        return

    # --- 阶段2: LLM 前预检。已存在的成本分析表只新增 BPM 实例，避免重复消耗 token。 ---
    creator_name = display_name or username
    bpm_no = normalize_bpm_no(bpm_no)
    llm_blocks: list[dict] = []
    reused_count = 0
    preflight_skipped = 0
    content_conflicts = 0

    for index, block in enumerate(blocks):
        quick_data = _quick_header_from_excel(saved_path, block)
        quick_code = (quick_data.get("quotation_no") or "").strip()
        content_hash = _block_content_hash(block)
        quote_date = _parse_date_value(quick_data.get("date_val"))
        if quick_code:
            existing = get_existing_quotation(db, tenant_id, quick_code)
            if existing:
                if getattr(existing, "content_hash", None) and content_hash and existing.content_hash != content_hash:
                    preflight_skipped += 1
                    content_conflicts += 1
                    yield {
                        "event": "skipped",
                        "quotation_code": quick_code,
                        "reason": "content_changed",
                        "message": f"同一成本分析号 {quick_code} 的表格内容与历史记录不一致，已跳过以避免覆盖旧数据",
                    }
                    continue
                try:
                    instance = ensure_bpm_instance(db, existing, bpm_no, quote_date or existing.analysis_date, creator_name, saved_path)
                except ValueError as exc:
                    preflight_skipped += 1
                    yield {
                        "event": "skipped",
                        "quotation_code": quick_code,
                        "reason": "quoted_instance",
                        "message": str(exc),
                    }
                    continue
                if content_hash and not getattr(existing, "content_hash", None):
                    existing.content_hash = content_hash
                db.flush()
                reused_count += 1
                yield {
                    "event": "reused",
                    "quotation_code": quick_code,
                    "bpm_no": bpm_no,
                    "instance_id": instance.id,
                    "message": f"复用已有成本分析表 {quick_code}，已登记 BPM流程号 {bpm_no}",
                }
                continue

        llm_blocks.append({
            "index": index,
            "block": block,
            "quick_data": quick_data,
            "content_hash": content_hash,
        })

    if reused_count or preflight_skipped:
        db.commit()

    if not llm_blocks:
        yield {
            "event": "complete",
            "processed": 0,
            "reused": reused_count,
            "total": total,
            "llm_errors": 0,
            "dup_within_batch": 0,
            "db_duplicates": 0,
            "content_conflicts": content_conflicts,
            "db_time": 0,
        }
        return

    # --- 阶段3: 并行 LLM 提取 ---
    semaphore = asyncio.Semaphore(settings.LLM_MAX_CONCURRENCY)
    results: list[dict] = []
    tasks: list[asyncio.Task] = []

    async def process_one(index: int, block: QuotationBlock, quick_data: dict, content_hash: str):
        t0 = time.time()
        try:
            data = await _extract_one_async(block, semaphore)
            t_llm = time.time() - t0
            code = data.get("quotation_no") or "?"
            logger.info(f"[{code}] LLM 完成, 耗时 {t_llm:.1f}s")
            return {
                "index": index,
                "block": block,
                "data": data,
                "error": None,
                "quotation_code": code,
                "llm_time": round(t_llm, 1),
                "quick_data": quick_data,
                "content_hash": content_hash,
            }
        except asyncio.CancelledError:
            logger.info(f"[Block {index}] LLM 任务被取消（客户端断开）")
            raise
        except Exception as e:
            t_llm = time.time() - t0
            logger.error(f"[Block {index}] LLM 失败 ({t_llm:.1f}s): {e}")
            return {
                "index": index,
                "block": block,
                "data": None,
                "error": str(e),
                "quotation_code": block.preview[:50],
                "llm_time": round(t_llm, 1),
                "quick_data": quick_data,
                "content_hash": content_hash,
            }

    try:
        tasks = [
            asyncio.create_task(process_one(item["index"], item["block"], item["quick_data"], item["content_hash"]))
            for item in llm_blocks
        ]

        completed = 0
        errors = 0
        for coro in asyncio.as_completed(tasks):
            result = await coro
            results.append(result)
            completed += 1
            if result["error"]:
                errors += 1
            yield {
                "event": "progress",
                "current": completed,
                "total": len(llm_blocks),
                "quotation_code": result["quotation_code"],
                "llm_time": result["llm_time"],
                "error": result["error"],
            }

        # --- 阶段3: 去重 + 写入数据库 ---
        results.sort(key=lambda r: r["index"])

        # 批次内去重：同一个成本分析号在一个 Excel 中出现多次，只保留第一个
        seen_codes: set[str] = set()
        dup_within_batch = 0
        deduped_results = []
        for r in results:
            if r["data"] is None:
                deduped_results.append(r)
                continue
            code = (r.get("quotation_code") or "").strip()
            if code and code in seen_codes:
                dup_within_batch += 1
                logger.warning(f"[{code}] 批次内重复，跳过")
                r["dup_skipped"] = True
                yield {
                    "event": "skipped",
                    "quotation_code": code,
                    "reason": "batch_dup",
                    "message": f"批次内重复：成本分析号 {code} 在本文件中已出现，保留第一个",
                }
            else:
                if code:
                    seen_codes.add(code)
                deduped_results.append(r)

        db_ready = sum(1 for r in deduped_results if r["data"] is not None and not r.get("dup_skipped"))
        yield {
            "event": "db_write",
            "message": f"LLM 解析完成，正在写入数据库 (有效 {db_ready}/{total}，批次内重复 {dup_within_batch})...",
        }

        t_db_start = time.time()
        db_success = 0
        db_duplicates = 0
        db_errors = 0

        for r in deduped_results:
            if r["data"] is None or r.get("dup_skipped"):
                continue
            try:
                data = _patch_header_fields_from_excel(r["data"], saved_path, r.get("block"))
                data = _patch_cost_summary_from_excel(data, saved_path, r.get("block"))
                final_code = (data.get("quotation_no") or r.get("quotation_code") or "").strip()
                content_hash = r.get("content_hash") or _block_content_hash(r.get("block"))
                existing = get_existing_quotation(db, tenant_id, final_code) if final_code else None
                if existing and getattr(existing, "content_hash", None) and content_hash and existing.content_hash != content_hash:
                    content_conflicts += 1
                    logger.warning(f"[{final_code}] 同号内容与历史不一致，跳过")
                    yield {
                        "event": "skipped",
                        "quotation_code": final_code,
                        "reason": "content_changed",
                        "message": f"同一成本分析号 {final_code} 的表格内容与历史记录不一致，已跳过以避免覆盖旧数据",
                    }
                    continue
                code = _write_one_quotation(
                    db,
                    data,
                    tenant_id,
                    username,
                    saved_path,
                    display_name,
                    bpm_no,
                    content_hash=content_hash,
                )
                if code:
                    db_success += 1
                    logger.info(f"[{code}] DB 写入完成")
                else:
                    db_duplicates += 1
                    reused_count += 1
                    logger.info(f"[{r['quotation_code']}] DB 中已存在，复用并登记 BPM 实例")
                    yield {
                        "event": "reused",
                        "quotation_code": r["quotation_code"],
                        "bpm_no": bpm_no,
                        "message": f"复用已有成本分析表 {r['quotation_code']}，已登记 BPM流程号 {bpm_no}",
                    }
            except ValueError as e:
                if "已报价" in str(e):
                    logger.warning(f"[{r['quotation_code']}] 已报价实例不允许覆盖: {e}")
                    db.rollback()
                    yield {
                        "event": "skipped",
                        "quotation_code": r["quotation_code"],
                        "reason": "quoted_instance",
                        "message": str(e),
                    }
                    continue
                logger.error(f"[{r['quotation_code']}] DB 写入失败: {e}")
                db.rollback()
                db_errors += 1
            except Exception as e:
                logger.error(f"[{r['quotation_code']}] DB 写入失败: {e}")
                db.rollback()
                db_errors += 1

        db.commit()
        t_db = time.time() - t_db_start

        total_skipped = dup_within_batch + content_conflicts
        logger.info(
            f"全部完成: LLM={completed}个 DB写入={db_success}个 "
            f"复用={reused_count}个 跳过={total_skipped}个(批内重复{dup_within_batch}/内容冲突{content_conflicts}) "
            f"DB耗时={t_db:.1f}s"
        )
        yield {
            "event": "complete",
            "processed": db_success,
            "reused": reused_count,
            "total": total,
            "llm_errors": errors,
            "dup_within_batch": dup_within_batch,
            "db_duplicates": db_duplicates,
            "content_conflicts": content_conflicts,
            "db_errors": db_errors,
            "db_time": round(t_db, 1),
        }

    except asyncio.CancelledError:
        logger.warning("ETL stream cancelled, cleaning up pending tasks")
        raise
    finally:
        cancelled_count = 0
        for t in tasks:
            if not t.done():
                t.cancel()
                cancelled_count += 1
        if cancelled_count > 0:
            logger.warning(f"Cancelled {cancelled_count} pending LLM tasks due to disconnect")


def get_upload_history(
    db: Session,
    tenant_id: str,
    creator_name: str,
    is_admin: bool = False,
    limit: int = 500,
    search: str = "",
):
    """查询报价单明细（按日期分组），管理员可查看所有。
    支持按成本分析号或 BPM 流程号搜索。"""
    from sqlalchemy import func, Date, or_

    search = (search or "").strip()

    filters = [
        QuotationMain.deleted == False,
        QuotationBpmInstance.deleted == False,
        QuotationBpmInstance.review_status != REVIEW_QUOTED,
    ]
    if not is_admin:
        filters.append(or_(QuotationMain.tenant_id == tenant_id, QuotationMain.tenant_id.is_(None)))
        admin_names = _admin_creator_names(db, tenant_id)
        if admin_names:
            filters.append(QuotationMain.creator.notin_(admin_names))

    if search:
        filters.append(or_(
            QuotationMain.quotation_code.contains(search),
            QuotationMain.bpm_no.contains(search),
            QuotationBpmInstance.bpm_no.contains(search),
            QuotationMain.customer_name.contains(search),
            QuotationMain.customer_address.contains(search),
            QuotationMain.package_method.contains(search),
            QuotationMain.product_spec.contains(search),
        ))

    rows = (
        db.query(
            QuotationMain.id,
            QuotationBpmInstance.id,
            QuotationMain.quotation_code,
            QuotationBpmInstance.bpm_no,
            QuotationMain.customer_name,
            QuotationMain.package_method,
            QuotationMain.product_spec,
            QuotationBpmInstance.upload_user,
            func.cast(QuotationBpmInstance.upload_time, Date).label("create_date"),
            QuotationBpmInstance.upload_time,
            QuotationBpmInstance.source_file_path,
            QuotationBpmInstance.review_status,
            QuotationBpmInstance.quote_date,
        )
        .join(QuotationMain, QuotationMain.id == QuotationBpmInstance.quotation_main_id)
        .filter(*filters)
        .order_by(QuotationBpmInstance.upload_time.desc(), QuotationBpmInstance.id.desc())
        .limit(limit)
        .all()
    )

    # 按日期分组
    from collections import OrderedDict
    groups = OrderedDict()
    instance_main_ids = set()
    for main_id, instance_id, code, bpm_no, customer, package_method, spec, creator, create_date, create_time, path, review_status, quote_date in rows:
        instance_main_ids.add(main_id)
        date_key = str(create_date) if create_date else "未知日期"
        if date_key not in groups:
            groups[date_key] = []
        groups[date_key].append({
            "instance_id": instance_id,
            "quotation_code": code,
            "bpm_no": bpm_no or "",
            "customer_name": customer or "",
            "package_method": package_method or "",
            "product_spec": spec or "",
            "upload_user": creator or "",
            "create_time": create_time.isoformat() if create_time else None,
            "quote_date": quote_date.isoformat() if quote_date else None,
            "filename": os.path.basename(path) if path else "",
            "review_status": review_status,
        })

    legacy_filters = [QuotationMain.deleted == False]
    if instance_main_ids:
        legacy_filters.append(QuotationMain.id.notin_(instance_main_ids))
    if not is_admin:
        legacy_filters.append(or_(QuotationMain.tenant_id == tenant_id, QuotationMain.tenant_id.is_(None)))
        admin_names = _admin_creator_names(db, tenant_id)
        if admin_names:
            legacy_filters.append(QuotationMain.creator.notin_(admin_names))
    if search:
        legacy_filters.append(or_(
            QuotationMain.quotation_code.contains(search),
            QuotationMain.bpm_no.contains(search),
            QuotationMain.customer_name.contains(search),
            QuotationMain.customer_address.contains(search),
            QuotationMain.package_method.contains(search),
            QuotationMain.product_spec.contains(search),
        ))
    legacy_rows = (
        db.query(
            QuotationMain.quotation_code,
            QuotationMain.bpm_no,
            QuotationMain.customer_name,
            QuotationMain.package_method,
            QuotationMain.product_spec,
            QuotationMain.creator,
            func.cast(QuotationMain.create_time, Date).label("create_date"),
            QuotationMain.create_time,
            QuotationMain.original_file_path,
            QuotationMain.extracted_tags,
            QuotationMain.analysis_date,
        )
        .filter(*legacy_filters)
        .order_by(QuotationMain.create_time.desc())
        .limit(limit)
        .all()
    )
    for code, bpm_no, customer, package_method, spec, creator, create_date, create_time, path, extracted_tags, quote_date in legacy_rows:
        from app.services.excel_preview_service import get_review_status_from_tags

        review_status = get_review_status_from_tags(extracted_tags)
        if review_status == REVIEW_QUOTED:
            continue
        date_key = str(create_date) if create_date else "未知日期"
        if date_key not in groups:
            groups[date_key] = []
        groups[date_key].append({
            "instance_id": None,
            "quotation_code": code,
            "bpm_no": bpm_no or "",
            "customer_name": customer or "",
            "package_method": package_method or "",
            "product_spec": spec or "",
            "upload_user": creator or "",
            "create_time": create_time.isoformat() if create_time else None,
            "quote_date": quote_date.isoformat() if quote_date else None,
            "filename": os.path.basename(path) if path else "",
            "review_status": review_status,
        })

    return [
        {"date": date, "items": items}
        for date, items in groups.items()
    ]


def delete_quotation(
    db: Session,
    quotation_code: str,
    tenant_id: str,
    creator_name: str,
    is_admin: bool = False,
    is_reviewer: bool = False,
    instance_id: int | None = None,
) -> bool:
    """删除指定成本分析号（管理员可删任意，普通用户仅可删自己），返回是否成功"""
    from sqlalchemy import or_

    if instance_id:
        query = (
            db.query(QuotationBpmInstance, QuotationMain)
            .join(QuotationMain, QuotationMain.id == QuotationBpmInstance.quotation_main_id)
            .filter(
                QuotationBpmInstance.id == instance_id,
                QuotationBpmInstance.deleted == False,
                QuotationMain.deleted == False,
            )
        )
        if not is_admin and not is_reviewer:
            query = query.filter(
                or_(QuotationMain.tenant_id == tenant_id, QuotationMain.tenant_id.is_(None)),
                QuotationBpmInstance.upload_user == creator_name,
            )
        row = query.first()
        if not row:
            return False
        instance, q = row
        if instance.review_status == REVIEW_QUOTED:
            return False
        instance.deleted = True
        instance.updater = creator_name
        instance.update_time = datetime.now()
        remaining = (
            db.query(QuotationBpmInstance)
            .filter(
                QuotationBpmInstance.quotation_main_id == q.id,
                QuotationBpmInstance.deleted == False,
                QuotationBpmInstance.id != instance.id,
            )
            .count()
        )
        if remaining == 0:
            _delete_quotation_fk_children(db, q.id)
            db.delete(q)
        db.commit()
        logger.info(f"[{quotation_code}] BPM实例 {instance_id} 已由 {creator_name} 删除 (admin={is_admin})")
        return True

    filters = [
        QuotationMain.quotation_code == quotation_code,
        QuotationMain.deleted == False,
    ]
    if not is_admin and not is_reviewer:
        filters.append(or_(QuotationMain.tenant_id == tenant_id, QuotationMain.tenant_id.is_(None)))
        filters.append(QuotationMain.creator == creator_name)

    q = db.query(QuotationMain).filter(*filters).first()
    if not q:
        return False
    from app.services.excel_preview_service import REVIEW_QUOTED, get_review_status
    if get_review_status(q) == REVIEW_QUOTED:
        return False
    _delete_quotation_fk_children(db, q.id)
    db.delete(q)
    db.commit()
    logger.info(f"[{quotation_code}] 已由 {creator_name} 删除 (admin={is_admin})")
    return True


# 保留同步版本，兼容旧的调用方式
def process_and_store_excel(file: UploadFile, db: Session, tenant_id: str, username: str, bpm_no: str = "") -> dict:
    """同步版：保存文件 + 扫描 + 串行 LLM + 写入（兼容旧接口）"""
    file_ext = os.path.splitext(file.filename)[1]
    unique_filename = f"{uuid.uuid4().hex}{file_ext}"
    saved_path = os.path.join(UPLOAD_DIR, unique_filename)

    with open(saved_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    logger.info(f"Saved original excel to {saved_path}")

    blocks = scan_quotations(saved_path)
    processed_count = 0
    seen_codes: set[str] = set()

    for block in blocks:
        try:
            t0 = time.time()
            data = asyncio.run(_extract_one_async(
                block, asyncio.Semaphore(settings.LLM_MAX_CONCURRENCY)
            ))
            t_llm = time.time() - t0

            # 批次内去重
            code = data.get("quotation_no") or f"AUTO-{uuid.uuid4().hex[:6]}"
            if code in seen_codes:
                logger.warning(f"[{code}] 批次内重复，跳过")
                continue
            seen_codes.add(code)

            t1 = time.time()
            data = _patch_header_fields_from_excel(data, saved_path, block)
            data = _patch_cost_summary_from_excel(data, saved_path, block)
            quotation_code = _write_one_quotation(
                db,
                data,
                tenant_id,
                username,
                saved_path,
                bpm_no=bpm_no,
                content_hash=_block_content_hash(block),
            )
            t_db = time.time() - t1

            if quotation_code:
                logger.info(f"[{quotation_code}] 耗时: LLM={t_llm:.1f}s  DB={t_db:.1f}s")
                db.commit()
                processed_count += 1
            else:
                logger.info(f"[{code}] 已存在，跳过")

        except Exception as e:
            db.rollback()
            import traceback
            logger.error(f"Error parsing quotation at row {block.row_idx}: {str(e)}\n{traceback.format_exc()}")

    return {"status": "success", "processed_quotations": processed_count, "saved_file": saved_path}
