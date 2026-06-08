from io import BytesIO
from types import SimpleNamespace

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.quotation import QuotationMain
from app.services.quotation_summary_service import apply_quotation_summaries


THIN_SIDE = Side(style="thin", color="000000")
CELL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
GRAY_FILL = PatternFill("solid", fgColor="F3F4F6")
TITLE_FILL = PatternFill("solid", fgColor="FFF900")
SUMMARY_FILL = PatternFill("solid", fgColor="FEF3C7")
FEE_LABEL_FILL = PatternFill("solid", fgColor="C6FFC6")
FEE_VALUE_FILL = PatternFill("solid", fgColor="F6A8D0")
PRICE_FILL = PatternFill("solid", fgColor="FFF900")


def render_quotation_excel(quotation: QuotationMain) -> BytesIO:
    """Generate an editable cost-analysis workbook from database fields."""
    apply_quotation_summaries(quotation)
    materials = sorted(
        (item for item in quotation.materials if not item.deleted),
        key=lambda item: (item.seq_no or 0, item.id or 0),
    )
    processes = sorted(
        (item for item in quotation.processes if not item.deleted),
        key=lambda item: item.id or 0,
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "成本分析表"
    worksheet.sheet_view.showGridLines = False

    widths = [10, 13, 18, 16, 16, 16, 16, 16, 18]
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width

    row = 1
    _merge_write(worksheet, row, 1, row, 9, "成 本 分 析 表", fill=TITLE_FILL, font=Font(name="宋体", size=20))
    worksheet.row_dimensions[row].height = 27

    row += 1
    _merge_write(worksheet, row, 1, row, 4, "")
    _write(worksheet, row, 5, "包装方式-米数:", bold=True)
    _merge_write(worksheet, row, 6, row, 7, getattr(quotation, "package_method", "") or "", horizontal="left")
    _write(worksheet, row, 8, "编号:", bold=True)
    _write(worksheet, row, 9, quotation.quotation_code)

    row += 1
    _write(worksheet, row, 1, "客户名称:", bold=True)
    _merge_write(worksheet, row, 2, row, 4, quotation.customer_name, horizontal="left")
    _write(worksheet, row, 5, "收货地（市）:", bold=True)
    _write(worksheet, row, 6, quotation.customer_address, horizontal="left")
    _write(worksheet, row, 7, "分析日期:", bold=True)
    _merge_write(worksheet, row, 8, row, 9, quotation.analysis_date, number_format="yyyy-mm-dd")

    row += 1
    _write(worksheet, row, 1, "结构:", bold=True)
    _merge_write(worksheet, row, 2, row, 4, quotation.structure, horizontal="left")
    _write(worksheet, row, 5, "编织率(%):", bold=True)
    _write(worksheet, row, 6, _percent_number(quotation.braiding_rate))
    _write(worksheet, row, 7, "品名规格:", bold=True)
    _merge_write(worksheet, row, 8, row, 9, quotation.product_spec, horizontal="left")

    row += 1
    material_header_row = row
    _merge_write(worksheet, row, 1, row + 1, 1, "序号", fill=GRAY_FILL, bold=True)
    _merge_write(worksheet, row, 2, row + 1, 2, "制程", fill=GRAY_FILL, bold=True)
    _merge_write(worksheet, row, 3, row + 1, 5, "规格", fill=GRAY_FILL, bold=True)
    _merge_write(worksheet, row, 6, row + 1, 6, "物料编码", fill=GRAY_FILL, bold=True)
    _write(worksheet, row, 7, "单位用量", fill=GRAY_FILL, bold=True)
    _write(worksheet, row, 8, "单价", fill=GRAY_FILL, bold=True)
    _write(worksheet, row, 9, "材料金额", fill=GRAY_FILL, bold=True)
    row += 1
    _write(worksheet, row, 7, "KG/100M", font=Font(size=9, color="6B7280"))
    _write(worksheet, row, 8, "RMB/KG", font=Font(size=9, color="6B7280"))
    _write(worksheet, row, 9, "RMB/100M", font=Font(size=9, color="6B7280"))

    for material in materials:
        row += 1
        _write(worksheet, row, 1, material.seq_no)
        _write(worksheet, row, 2, material.process_name)
        _merge_write(worksheet, row, 3, row, 5, material.spec_detail, horizontal="left")
        _write(worksheet, row, 6, material.process_code)
        _write(worksheet, row, 7, material.unit_usage)
        _write(worksheet, row, 8, material.unit_price)
        _write(worksheet, row, 9, material.material_amount)
    if not materials:
        row += 1
        _merge_write(worksheet, row, 1, row, 9, "暂无材料明细")

    row += 1
    _merge_write(worksheet, row, 1, row, 2, "材料成本", fill=SUMMARY_FILL, bold=True)
    _write(worksheet, row, 3, quotation.material_cost, fill=SUMMARY_FILL, bold=True)
    _merge_write(worksheet, row, 4, row, 6, "RMB/M", font=Font(size=9, color="6B7280"))
    _write(worksheet, row, 7, quotation.unit_usage_sum, fill=SUMMARY_FILL, bold=True)
    _write(worksheet, row, 8, "Kg", font=Font(size=9, color="6B7280"))
    _write(worksheet, row, 9, quotation.material_amount_sum, fill=SUMMARY_FILL, bold=True)

    row += 1
    process_header_row = row
    process_headers = [
        "制程",
        "标准工时(一台机1KM开机时间)",
        "损耗时间(1KM)",
        "固定费用率",
        "固定费用",
        "开机损耗废线",
        "每个制程总废胶(KG)",
        "金额",
        "费用成本小计",
    ]
    for column, header in enumerate(process_headers, start=1):
        _write(worksheet, row, column, header, fill=GRAY_FILL, bold=True, wrap_text=True)
    worksheet.row_dimensions[row].height = 32

    for process in processes:
        row += 1
        values = [
            process.process_name,
            process.std_hours,
            process.loss_hours,
            process.fixed_rate,
            process.fixed_fee,
            process.startup_loss_wire,
            process.total_waste_glue,
            process.amount,
            process.subtotal_fee,
        ]
        for column, value in enumerate(values, start=1):
            _write(worksheet, row, column, value)
    if not processes:
        row += 1
        _merge_write(worksheet, row, 1, row, 9, "暂无制程费用明细")

    row += 1
    fee_start_row = row
    _merge_write(worksheet, row, 1, row + 5, 2, "其他费用", fill=FEE_LABEL_FILL, bold=True)
    fee_headers = ["UL标签费(RMB/M)", "运输费(RMB/KG)", "包装费(RMB/M)", "废品损耗(%)", "订单开机次数", "费用总计"]
    for column, header in enumerate(fee_headers, start=3):
        _write(worksheet, row, column, header, fill=FEE_LABEL_FILL, bold=True, wrap_text=True)
    _write(worksheet, row, 9, quotation.total_fee, fill=PRICE_FILL, font=Font(color="FF0000"))

    row += 1
    _write(worksheet, row, 3, quotation.ul_label_fee, fill=FEE_VALUE_FILL)
    _write(worksheet, row, 4, quotation.transport_fee, fill=FEE_VALUE_FILL)
    _write(worksheet, row, 5, quotation.packing_fee, fill=FEE_VALUE_FILL)
    _write(worksheet, row, 6, quotation.waste_loss_rate, fill=FEE_VALUE_FILL, number_format="0.####%")
    _write(worksheet, row, 7, quotation.order_startup_times, fill=FEE_VALUE_FILL)
    _write(worksheet, row, 8, "成本(RMB/M)", fill=PRICE_FILL, bold=True)
    _write(worksheet, row, 9, "取利售价(RMB/M)", fill=PRICE_FILL, bold=True)

    row += 1
    fee_headers = ["其他费用", "净利率", "报关费(RMB/次)", "增值税率", "订单米数"]
    for column, header in enumerate(fee_headers, start=3):
        _write(worksheet, row, column, header, fill=FEE_LABEL_FILL, bold=True)
    _write(worksheet, row, 8, quotation.cost, fill=PRICE_FILL, font=Font(color="FF0000"))
    _write(worksheet, row, 9, quotation.profit_selling_price, fill=PRICE_FILL, font=Font(color="0000D7", bold=True, size=13))

    row += 1
    _write(worksheet, row, 3, quotation.other_fee, fill=FEE_VALUE_FILL)
    _write(worksheet, row, 4, quotation.net_profit_rate, fill=FEE_VALUE_FILL, number_format="0.####%")
    _write(worksheet, row, 5, quotation.customs_fee, fill=FEE_VALUE_FILL)
    _write(worksheet, row, 6, quotation.vat_rate, fill=FEE_VALUE_FILL, number_format="0.####%")
    _write(worksheet, row, 7, quotation.order_meterage, fill=FEE_VALUE_FILL)
    _write(worksheet, row, 8, "不取利售价(RMB/M)", fill=PRICE_FILL, bold=True)
    _write(worksheet, row, 9, "最终售价(RMB/M)", fill=PRICE_FILL, bold=True)

    row += 1
    fee_headers = ["照射芯数", "照射费用(RMB/M)", "营业费用率", "月结利息", "企税税率"]
    for column, header in enumerate(fee_headers, start=3):
        _write(worksheet, row, column, header, fill=FEE_LABEL_FILL, bold=True)
    _write(worksheet, row, 8, quotation.non_profit_price, fill=PRICE_FILL, font=Font(color="FF0000"))
    _write(worksheet, row, 9, quotation.final_selling_price, fill=PRICE_FILL, font=Font(color="0000D7", bold=True, size=13))

    row += 1
    _write(worksheet, row, 3, quotation.irradiation_core_count, fill=FEE_VALUE_FILL)
    _write(worksheet, row, 4, quotation.irradiation_core_fee, fill=FEE_VALUE_FILL)
    _write(worksheet, row, 5, quotation.operating_expense_rate, fill=FEE_VALUE_FILL, number_format="0.####%")
    _write(worksheet, row, 6, quotation.monthly_interest, fill=FEE_VALUE_FILL, number_format="0.####%")
    _write(worksheet, row, 7, quotation.corporate_tax_rate, fill=FEE_VALUE_FILL, number_format="0.####%")
    _merge_write(worksheet, row, 8, row, 9, "", fill=PRICE_FILL)

    row += 1
    _write(worksheet, row, 1, "备注:", bold=True)
    _merge_write(worksheet, row, 2, row, 9, quotation.remark or "", horizontal="left", wrap_text=True)

    worksheet.freeze_panes = f"A{material_header_row + 2}"
    worksheet.print_area = f"A1:I{row}"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.print_title_rows = f"1:{material_header_row + 1}"
    _ = process_header_row, fee_start_row

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def render_quote_snapshot_excel(snapshot_data: dict) -> BytesIO:
    main = snapshot_data.get("main") or {}
    materials = [
        SimpleNamespace(**{**item, "deleted": False})
        for item in (snapshot_data.get("materials") or [])
    ]
    processes = [
        SimpleNamespace(**{**item, "deleted": False})
        for item in (snapshot_data.get("processes") or [])
    ]
    quotation = SimpleNamespace(
        **main,
        materials=materials,
        processes=processes,
        deleted=False,
        updater="",
        update_time=None,
    )
    return render_quotation_excel(quotation)


def _write(
    worksheet,
    row: int,
    column: int,
    value="",
    *,
    fill=WHITE_FILL,
    font=None,
    bold: bool = False,
    horizontal: str = "center",
    wrap_text: bool = False,
    number_format: str | None = None,
):
    cell = worksheet.cell(row=row, column=column, value=_excel_value(value))
    cell.border = CELL_BORDER
    cell.fill = fill
    cell.font = font or Font(bold=bold)
    cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap_text)
    if number_format:
        cell.number_format = number_format
    return cell


def _merge_write(worksheet, min_row: int, min_col: int, max_row: int, max_col: int, value="", **kwargs):
    worksheet.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
    cell = _write(worksheet, min_row, min_col, value, **kwargs)
    for row in range(min_row, max_row + 1):
        for column in range(min_col, max_col + 1):
            merged_cell = worksheet.cell(row=row, column=column)
            merged_cell.border = CELL_BORDER
            merged_cell.fill = kwargs.get("fill", WHITE_FILL)
    return cell


def _excel_value(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def _percent_number(value):
    if value in (None, ""):
        return ""
    return _excel_value(value) * 100
