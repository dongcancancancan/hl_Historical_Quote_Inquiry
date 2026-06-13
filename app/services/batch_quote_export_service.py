from __future__ import annotations

import re
import os
from copy import copy
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from sqlalchemy.orm import Session

from app.models.calc_param import QuotationCalcParam
from app.models.quotation import QuotationBpmInstance, QuotationMain
from app.services.bpm_instance_service import REVIEW_PARAM_FIELDS
from app.services.calc_param_service import DEFAULT_COPPER_ROD_PROCESS_FEE, DEFAULT_VAT_RATE, normalize_vat_multiplier, normalize_vat_rate
from app.services.copper_scenario_service import _calculate_one_band

PROJECT_ROOT = Path(__file__).resolve().parents[2]
GENERAL_TEMPLATE_PATH = PROJECT_ROOT / "通用报价单格式.xlsx"
GENERAL_TEMPLATE_SHEET = "2000"
GENERAL_TEMPLATE_ID = "general_quote_xls"

QUOTE_TEMPLATES = {
    GENERAL_TEMPLATE_ID: {
        "id": GENERAL_TEMPLATE_ID,
        "name": "通用报价单格式",
        "filename": GENERAL_TEMPLATE_PATH.name,
        "sheet_name": GENERAL_TEMPLATE_SHEET,
        "description": "鸿林通用报价单模板",
    }
}

TEMPLATE_BANDS = [
    {"label": "90001-92000", "header": "含税铜价\n90001-92000", "copper_price": Decimal("92000")},
    {"label": "92001-94000", "header": "含税铜价\n92001-94000", "copper_price": Decimal("94000")},
    {"label": "94001-96000", "header": "含税铜价\n94001-96000", "copper_price": Decimal("96000")},
    {"label": "96001-98000", "header": "含税铜价\n96001-98000", "copper_price": Decimal("98000")},
    {"label": "98001-100000", "header": "含税铜价\n98001-100000", "copper_price": Decimal("100000")},
    {"label": "100001-102000", "header": "含税铜价\n100001-102000", "copper_price": Decimal("102000")},
    {"label": "102001-104000", "header": "含税铜价\n102001-104000", "copper_price": Decimal("104000")},
    {"label": "104001-106000", "header": "含税铜价\n104001-106000", "copper_price": Decimal("106000")},
    {"label": "106001-108000", "header": "含税铜价\n106001-108000", "copper_price": Decimal("108000")},
    {"label": "108001-110000", "header": "含税铜价\n108001-110000", "copper_price": Decimal("110000")},
]


def render_general_batch_quote_excel(
    db: Session,
    instance_ids: list[int],
    operator: str,
    template_id: str = GENERAL_TEMPLATE_ID,
) -> tuple[BytesIO, str]:
    template = _get_quote_template(template_id)
    ordered_ids = _normalize_instance_ids(instance_ids)
    if not ordered_ids:
        raise ValueError("请先选择需要导出报价单的 BPM 实例")

    rows = (
        db.query(QuotationBpmInstance, QuotationMain)
        .join(QuotationMain, QuotationMain.id == QuotationBpmInstance.quotation_main_id)
        .filter(
            QuotationBpmInstance.id.in_(ordered_ids),
            QuotationBpmInstance.deleted == False,
            QuotationMain.deleted == False,
        )
        .all()
    )
    row_map = {instance.id: (instance, quotation) for instance, quotation in rows}
    missing_ids = [item_id for item_id in ordered_ids if item_id not in row_map]
    if missing_ids:
        raise ValueError(f"以下 BPM 实例不存在或已删除：{', '.join(str(item) for item in missing_ids)}")

    export_rows = []
    bpm_values: list[str] = []
    customer_names: list[str] = []
    customer_addresses: list[str] = []
    for item_id in ordered_ids:
        instance, quotation = row_map[item_id]
        current_final = instance.final_selling_price if instance.final_selling_price is not None else quotation.final_selling_price
        if current_final in (None, ""):
            raise ValueError(
                f"{quotation.quotation_code or item_id} 尚未生成当前最终售价，请先在批量页执行“保存参数并一键计算”"
            )
        export_rows.append(_build_export_row(db, quotation, instance))
        if instance.bpm_no:
            bpm_values.append(instance.bpm_no)
        if quotation.customer_name:
            customer_names.append(quotation.customer_name)
        if quotation.customer_address:
            customer_addresses.append(quotation.customer_address)

    bpm_no = _shared_text(bpm_values)
    customer_name = _shared_text(customer_names)
    customer_address = _shared_text(customer_addresses)
    buffer = _render_workbook_from_template(
        template=template,
        rows=export_rows,
        bpm_no=bpm_no,
        customer_name=customer_name,
        customer_address=customer_address,
        operator=operator,
    )
    bpm_for_name = _shared_text(bpm_values) or datetime.now().strftime("%Y%m%d%H%M%S")
    return buffer, f"{bpm_for_name}-报价单.xlsx"


def list_quote_templates() -> list[dict]:
    return [
        {
            "id": item["id"],
            "name": item["name"],
            "filename": item["filename"],
            "description": item["description"],
        }
        for item in QUOTE_TEMPLATES.values()
    ]


def _get_quote_template(template_id: str | None) -> dict:
    selected_id = template_id or GENERAL_TEMPLATE_ID
    template = QUOTE_TEMPLATES.get(selected_id)
    if not template:
        raise ValueError(f"未知报价单模板：{selected_id}")
    return template


def _build_export_row(db: Session, quotation: QuotationMain, instance: QuotationBpmInstance) -> dict:
    band_results = _calculate_template_bands(db, quotation, instance)
    jacket_row = _find_export_material(quotation)
    return {
        "quotation_code": quotation.quotation_code or "",
        "product_spec": quotation.product_spec or "",
        "product_name": _build_product_name(quotation, jacket_row),
        "band_prices": [item["final_selling_price"] for item in band_results],
    }


def _calculate_template_bands(db: Session, quotation: QuotationMain, instance: QuotationBpmInstance) -> list[dict]:
    params = (
        db.query(QuotationCalcParam)
        .filter(QuotationCalcParam.quotation_main_id == quotation.id)
        .first()
    )
    simulated_params = SimpleNamespace(
        copper_rod_process_fee=instance.copper_rod_process_fee
        or (params.copper_rod_process_fee if params else DEFAULT_COPPER_ROD_PROCESS_FEE),
        vat_rate=normalize_vat_multiplier(params.vat_rate if params else DEFAULT_VAT_RATE),
    )
    original_values = _snapshot_review_values(quotation)
    try:
        _apply_instance_review_values(quotation, instance)
        results = []
        for band in TEMPLATE_BANDS:
            result = _calculate_one_band(db, quotation, simulated_params, band["copper_price"])
            if not result.get("final_selling_price"):
                raise ValueError(f"{quotation.quotation_code or quotation.id} 在铜段 {band['label']} 未生成最终售价")
            results.append(result)
        return results
    finally:
        _restore_review_values(quotation, original_values)


def _snapshot_review_values(quotation: QuotationMain) -> dict:
    values = {field: getattr(quotation, field, None) for field in REVIEW_PARAM_FIELDS}
    values["vat_rate"] = getattr(quotation, "vat_rate", None)
    return values


def _apply_instance_review_values(quotation: QuotationMain, instance: QuotationBpmInstance) -> None:
    for field in REVIEW_PARAM_FIELDS:
        value = getattr(instance, field, None)
        if value is not None:
            setattr(quotation, field, value)
    if instance.vat_rate is not None:
        quotation.vat_rate = normalize_vat_rate(instance.vat_rate)


def _restore_review_values(quotation: QuotationMain, values: dict) -> None:
    for field, value in values.items():
        setattr(quotation, field, value)


def _find_export_material(quotation: QuotationMain):
    materials = [item for item in quotation.materials if not item.deleted]
    if not materials:
        return None
    materials.sort(key=lambda item: (item.seq_no if item.seq_no is not None else 10**9, item.id or 0))
    return materials[-1]


def _build_product_name(quotation: QuotationMain, jacket_row) -> str:
    parts = []
    if quotation.structure:
        parts.append(str(quotation.structure).strip())
    material_name = _outer_material_name(jacket_row)
    if material_name:
        parts.append(material_name)
    od_or_id = _extract_od_or_id(quotation, jacket_row)
    if od_or_id:
        parts.append(od_or_id)
    return " ".join(part for part in parts if part).strip() or (quotation.product_spec or quotation.quotation_code or "")


def _outer_material_name(jacket_row) -> str:
    if not jacket_row:
        return ""
    code = str(getattr(jacket_row, "process_code", "") or "").strip().upper()
    if code.startswith("EX"):
        return "XLPE外被"
    if code.startswith("EE"):
        return "TPE外被"
    if code.startswith("C"):
        return "低毒PVC外被"
    name = str(getattr(jacket_row, "process_name", "") or "").strip()
    return name or "外被"


def _extract_od_or_id(quotation: QuotationMain, jacket_row) -> str:
    candidates = [
        getattr(jacket_row, "spec_detail", "") if jacket_row else "",
        getattr(jacket_row, "process_name", "") if jacket_row else "",
        quotation.product_spec or "",
        quotation.structure or "",
        quotation.remark or "",
    ]
    pattern = re.compile(r"\b(OD|ID)\s*[:：]?\s*([0-9]+(?:\.[0-9]+)?)\s*mm\b", re.IGNORECASE)
    for text in candidates:
        match = pattern.search(str(text or ""))
        if match:
            return f"{match.group(1).upper()}：{match.group(2)} mm"
    return ""


def _template_path(template: dict) -> Path:
    candidates = _template_path_candidates(template)
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def _template_path_candidates(template: dict) -> list[Path]:
    filename = template["filename"]
    explicit_path = os.getenv("QUOTE_TEMPLATE_PATH")
    explicit_dir = os.getenv("QUOTE_TEMPLATE_DIR")
    candidates = []
    if explicit_path:
        candidates.append(Path(explicit_path))
    if explicit_dir:
        candidates.append(Path(explicit_dir) / filename)
    candidates.extend(
        [
            PROJECT_ROOT / filename,
            PROJECT_ROOT / "app" / filename,
            Path.cwd() / filename,
            Path.cwd() / "app" / filename,
            Path("/app") / filename,
            Path("/app/app") / filename,
            Path("/data/hl_Historical_Quote_Inquiry") / filename,
            Path("/data/hl_Historical_Quote_Inquiry/app") / filename,
            Path("/data") / filename,
        ]
    )
    return list(dict.fromkeys(candidates))


def _render_workbook_from_template(
    template: dict,
    rows: list[dict],
    bpm_no: str,
    customer_name: str,
    customer_address: str,
    operator: str,
) -> BytesIO:
    template_path = _template_path(template)
    if not template_path.exists():
        searched = "；".join(str(item) for item in _template_path_candidates(template))
        raise ValueError(f"未找到报价单模板：{template_path}；已搜索路径：{searched}")

    try:
        workbook = load_workbook(template_path)
    except Exception as exc:
        raise ValueError(f"无法打开报价单模板：{template_path}（{exc}）") from exc
    if template["sheet_name"] not in workbook.sheetnames:
        raise ValueError(f"报价单模板中未找到工作表：{template['sheet_name']}")
    sheet = workbook[template["sheet_name"]]
    _fill_template_sheet(sheet, rows, bpm_no, customer_name, customer_address, operator)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _fill_template_sheet(
    sheet,
    rows: list[dict],
    bpm_no: str,
    customer_name: str,
    customer_address: str,
    operator: str,
) -> None:
    detail_start_row = 14
    reserved_detail_rows = 4
    detail_count = max(len(rows), reserved_detail_rows)
    extra_rows = max(0, detail_count - reserved_detail_rows)
    if extra_rows:
        insert_at = detail_start_row + reserved_detail_rows
        sheet.insert_rows(insert_at, extra_rows)
        _shift_images_below(sheet, insert_at, extra_rows)
        source_row = insert_at - 1
        for offset in range(extra_rows):
            _copy_row_format(sheet, source_row, insert_at + offset, max_col=13)

    # 解除数据行区域内的合并单元格，避免铜段价格写入时互相覆盖
    first_data_row = detail_start_row
    last_data_row = detail_start_row + detail_count - 1
    _unmerge_data_area(sheet, first_data_row, last_data_row)

    _set_cell_value(sheet, 7, 2, customer_name or "")
    _set_cell_value(sheet, 8, 2, customer_address or "")
    _set_cell_value(sheet, 10, 4, operator or "")
    _set_cell_value(sheet, 11, 2, bpm_no or "")
    _set_cell_value(sheet, 11, 9, f"日期：{datetime.now():%Y-%m-%d}")

    for offset in range(detail_count):
        row = rows[offset] if offset < len(rows) else None
        _write_template_detail_row(sheet, detail_start_row + offset, row)



def _write_template_detail_row(sheet, row_no: int, row: dict | None) -> None:
    values = ["", "", "", *[""] * 10]
    if row:
        values[0] = row["product_name"]
        values[1] = row["product_spec"]
        values[2] = row["quotation_code"]
        for idx, price in enumerate(row["band_prices"], start=3):
            values[idx] = _excel_number(price)
    for index, value in enumerate(values, start=1):
        _set_cell_value(sheet, row_no, index, value)


def _set_cell_value(sheet, row_no: int, col_no: int, value) -> None:
    cell = sheet.cell(row=row_no, column=col_no)
    if isinstance(cell, MergedCell):
        cell = _merged_range_anchor(sheet, row_no, col_no)
    cell.value = value


def _merged_range_anchor(sheet, row_no: int, col_no: int):
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row <= row_no <= merged_range.max_row and merged_range.min_col <= col_no <= merged_range.max_col:
            return sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
    return sheet.cell(row=row_no, column=col_no)


def _unmerge_data_area(sheet, first_row: int, last_row: int) -> None:
    """移除数据行区域内的所有合并单元格。"""
    to_remove = []
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row <= last_row and merged_range.max_row >= first_row:
            to_remove.append(str(merged_range))
    for ref in to_remove:
        sheet.unmerge_cells(ref)


def _copy_row_format(sheet, source_row: int, target_row: int, max_col: int) -> None:
    source_dimension = sheet.row_dimensions[source_row]
    target_dimension = sheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    for col_no in range(1, max_col + 1):
        source_cell = sheet.cell(row=source_row, column=col_no)
        target_cell = sheet.cell(row=target_row, column=col_no)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)


def _shift_images_below(sheet, insert_at_row: int, amount: int) -> None:
    threshold = insert_at_row - 1
    for image in getattr(sheet, "_images", []):
        anchor = getattr(image, "anchor", None)
        for attr in ("_from", "_to", "to"):
            marker = getattr(anchor, attr, None)
            if marker is not None and getattr(marker, "row", -1) >= threshold:
                marker.row += amount


def _normalize_instance_ids(instance_ids: list[int]) -> list[int]:
    ordered = []
    seen = set()
    for value in instance_ids or []:
        try:
            item = int(value)
        except (TypeError, ValueError):
            continue
        if item <= 0 or item in seen:
            continue
        seen.add(item)
        ordered.append(item)
    return ordered


def _shared_text(values: list[str]) -> str:
    cleaned = [str(value).strip() for value in values if str(value or "").strip()]
    if not cleaned:
        return ""
    unique = []
    seen = set()
    for item in cleaned:
        if item in seen:
            continue
        seen.add(item)
        unique.append(item)
    return unique[0] if len(unique) == 1 else " / ".join(unique)


def _excel_number(value):
    if value in (None, ""):
        return ""
    try:
        return float(Decimal(str(value)))
    except Exception:
        return value
