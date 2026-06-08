import json
import re
from datetime import datetime
from decimal import Decimal

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.copper_fee import CopperProcessingFee, CopperProcessingFeeLog


DEFAULT_TC_TIN_PRICE_BASIS = Decimal("350")
COPPER_CODE_RE = re.compile(r"^\s*(\d+(?:\.\d+)?)\s*(BC|TC|TD)\s*$", re.IGNORECASE)


def normalize_copper_type(value: str) -> str:
    copper_type = (value or "").strip().upper()
    if copper_type == "TD":
        copper_type = "TC"  # TD 视为 TC（镀锡铜）
    if copper_type not in {"BC", "TC"}:
        raise ValueError("铜类型仅支持 BC、TC 或 TD")
    return copper_type


def normalize_tin_price_basis(copper_type: str, value=None) -> Decimal:
    # TD 已归一化为 TC，这里按 TC 处理
    if copper_type == "BC":
        return Decimal("0")
    if value in (None, ""):
        return DEFAULT_TC_TIN_PRICE_BASIS
    return Decimal(str(value))


def parse_copper_material_code(material_code: str) -> dict:
    match = COPPER_CODE_RE.match(material_code or "")
    if not match:
        raise ValueError("物料编码格式应为线径加 BC/TC/TD，例如 0.196BC")
    copper_type = match.group(2).upper()
    if copper_type == "TD":
        copper_type = "TC"
    return {
        "diameter": Decimal(match.group(1)),
        "copper_type": copper_type,
        "tin_price_basis": normalize_tin_price_basis(copper_type),
    }


def match_copper_processing_fee(db: Session, material_code: str, tin_price_basis=None):
    parsed = parse_copper_material_code(material_code)
    basis = normalize_tin_price_basis(parsed["copper_type"], tin_price_basis)
    fee = db.query(CopperProcessingFee).filter(
        CopperProcessingFee.copper_type == parsed["copper_type"],
        CopperProcessingFee.diameter == parsed["diameter"],
        CopperProcessingFee.tin_price_basis == basis,
        CopperProcessingFee.enabled == True,
    ).first()
    return parsed, fee


def list_copper_fees(db: Session, copper_type: str = "", keyword: str = "", include_disabled: bool = False):
    query = db.query(CopperProcessingFee)
    if copper_type:
        query = query.filter(CopperProcessingFee.copper_type == normalize_copper_type(copper_type))
    if keyword:
        try:
            query = query.filter(CopperProcessingFee.diameter == Decimal(keyword))
        except Exception:
            return []
    if not include_disabled:
        query = query.filter(CopperProcessingFee.enabled == True)
    return query.order_by(CopperProcessingFee.copper_type, CopperProcessingFee.tin_price_basis, CopperProcessingFee.diameter).all()


def create_copper_fee(db: Session, data: dict, operator: str, action: str = "CREATE", commit: bool = True):
    normalized = _normalize_payload(data)
    existing = _find_fee(db, normalized)
    if existing:
        raise ValueError("相同类型、线径和锡价段的加工费已存在")
    fee = CopperProcessingFee(**normalized, creator=operator, updater=operator, update_time=datetime.now())
    db.add(fee)
    db.flush()
    _write_log(db, fee.id, action, None, _snapshot(fee), operator)
    if commit:
        db.commit()
    return fee


def update_copper_fee(db: Session, fee: CopperProcessingFee, data: dict, operator: str, action: str = "UPDATE", commit: bool = True):
    before = _snapshot(fee)
    normalized = _normalize_payload(data, fee)
    duplicate = _find_fee(db, normalized, exclude_id=fee.id)
    if duplicate:
        raise ValueError("相同类型、线径和锡价段的加工费已存在")
    for key, value in normalized.items():
        setattr(fee, key, value)
    fee.updater = operator
    fee.update_time = datetime.now()
    _write_log(db, fee.id, action, before, _snapshot(fee), operator)
    if commit:
        db.commit()
    return fee


def disable_copper_fee(db: Session, fee: CopperProcessingFee, operator: str):
    before = _snapshot(fee)
    fee.enabled = False
    fee.updater = operator
    fee.update_time = datetime.now()
    _write_log(db, fee.id, "DELETE", before, _snapshot(fee), operator)
    db.commit()


def import_copper_fees(db: Session, file_path: str, operator: str) -> dict:
    workbook = load_workbook(file_path, data_only=True)
    worksheet = workbook.active
    created = 0
    updated = 0
    try:
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            _, copper_type, diameter, processing_fee, _, minimum_fee = (list(row) + [None] * 6)[:6]
            if not copper_type or diameter in (None, "") or processing_fee in (None, ""):
                continue
            payload = {
                "copper_type": copper_type,
                "diameter": diameter,
                "tin_price_basis": DEFAULT_TC_TIN_PRICE_BASIS if str(copper_type).upper() == "TC" else 0,
                "processing_fee": processing_fee,
                "minimum_fee": minimum_fee,
                "enabled": True,
            }
            normalized = _normalize_payload(payload)
            existing = _find_fee(db, normalized)
            if existing:
                update_copper_fee(db, existing, payload, operator, action="IMPORT", commit=False)
                updated += 1
            else:
                create_copper_fee(db, payload, operator, action="IMPORT", commit=False)
                created += 1
    finally:
        workbook.close()
    db.commit()
    return {"created": created, "updated": updated}


def serialize_fee(fee: CopperProcessingFee) -> dict:
    return {
        "id": fee.id,
        "copper_type": fee.copper_type,
        "diameter": _decimal_text(fee.diameter),
        "tin_price_basis": _decimal_text(fee.tin_price_basis),
        "processing_fee": _decimal_text(fee.processing_fee),
        "minimum_fee": _decimal_text(fee.minimum_fee),
        "remark": fee.remark or "",
        "enabled": bool(fee.enabled),
        "creator": fee.creator or "",
        "create_time": fee.create_time.isoformat() if fee.create_time else None,
        "updater": fee.updater or "",
        "update_time": fee.update_time.isoformat() if fee.update_time else None,
    }


def serialize_log(log: CopperProcessingFeeLog) -> dict:
    return {
        "id": log.id,
        "fee_id": log.fee_id,
        "action": log.action,
        "before_data": json.loads(log.before_data) if log.before_data else None,
        "after_data": json.loads(log.after_data) if log.after_data else None,
        "operator": log.operator,
        "operate_time": log.operate_time.isoformat() if log.operate_time else None,
    }


def _normalize_payload(data: dict, existing=None) -> dict:
    copper_type = normalize_copper_type(data.get("copper_type", existing.copper_type if existing else ""))
    diameter = Decimal(str(data.get("diameter", existing.diameter if existing else "")))
    processing_fee = Decimal(str(data.get("processing_fee", existing.processing_fee if existing else "")))
    minimum_raw = data.get("minimum_fee", existing.minimum_fee if existing else None)
    minimum_fee = None if minimum_raw in (None, "") else Decimal(str(minimum_raw))
    return {
        "copper_type": copper_type,
        "diameter": diameter,
        "tin_price_basis": normalize_tin_price_basis(copper_type, data.get("tin_price_basis", existing.tin_price_basis if existing else None)),
        "processing_fee": processing_fee,
        "minimum_fee": minimum_fee,
        "remark": str(data.get("remark", existing.remark if existing else "") or "").strip(),
        "enabled": bool(data.get("enabled", existing.enabled if existing else True)),
    }


def _find_fee(db: Session, data: dict, exclude_id: int | None = None):
    query = db.query(CopperProcessingFee).filter(
        CopperProcessingFee.copper_type == data["copper_type"],
        CopperProcessingFee.diameter == data["diameter"],
        CopperProcessingFee.tin_price_basis == data["tin_price_basis"],
    )
    if exclude_id:
        query = query.filter(CopperProcessingFee.id != exclude_id)
    return query.first()


def _write_log(db: Session, fee_id: int, action: str, before: dict | None, after: dict | None, operator: str):
    db.add(CopperProcessingFeeLog(
        fee_id=fee_id,
        action=action,
        before_data=json.dumps(before, ensure_ascii=False) if before else None,
        after_data=json.dumps(after, ensure_ascii=False) if after else None,
        operator=operator,
    ))


def _snapshot(fee: CopperProcessingFee) -> dict:
    return serialize_fee(fee)


def _decimal_text(value) -> str | None:
    if value is None:
        return None
    return f"{Decimal(value):f}".rstrip("0").rstrip(".") or "0"
